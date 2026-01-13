
# # import requests

# # def download_page(url):
# #     headers = {'User-Agent': 'Mozilla/5.0'}  # 模拟浏览器
# #     try:
# #         response = requests.get(url, headers=headers)
# #         response.raise_for_status()  # 检查HTTP错误
# #         return response.text
# #     except requests.exceptions.RequestException as e:
# #         print(f"下载失败: {e}")
# #         return None

# # url = 'https://jira.sw.nxp.com/sr/jira.issueviews:searchrequest-csv-with-bom-all-fields/temp/SearchRequest.csv?jqlQuery=project+%3D+ACSVS+AND+affectedVersion+%3D+i.mxRT2660+ORDER+BY+priority+DESC%2C+updated+DESC'
# # content = download_page(url)
# # if content:
# #     print(content)


# # https://jira.sw.nxp.com/sr/jira.issueviews:searchrequest-csv-with-bom-all-fields/temp/SearchRequest.csv?jqlQuery=project+%3D+ACSVS+AND+affectedVersion+%3D+i.mxRT2660+ORDER+BY+priority+DESC%2C+updated+DESC

# # import requests
# # from requests.auth import HTTPBasicAuth
# # import os
# # url = 'https://jira.sw.nxp.com/sr/jira.issueviews:searchrequest-csv-with-bom-all-fields/temp/SearchRequest.csv?jqlQuery=project+%3D+ACSVS+AND+affectedVersion+%3D+i.mxRT2660+ORDER+BY+priority+DESC%2C+updated+DESC'
# # path = r'C:\Users\nxa28190\OneDrive - NXP\test_code\python\projects\python3\jira_ticket\test.csv'

# # def download_excel():
# 	# current_path = os.getcwd()
# 	# original_csv = os.path.join(os.getcwd(), 'original_csv.csv')
# 	# print(original_csv)

# 	# # 使用os.path.exists()检查文件是否存在
# 	# if not os.path.exists(original_csv):
# 	#     # 如果文件不存在，使用open()创建文件
# 	#     open(original_csv, 'w').close()  # 使用'w'模式打开文件，这将创建文件，如果文件已存在则覆盖
# 	#     print("文件已创建")
# 	# else:
# 	#     print("文件已存在")


# 	# # 配置信息（需替换为你的实际数据）
# 	# url = "https://jira.sw.nxp.com/sr/jira.issueviews:searchrequest-csv-with-bom-all-fields/temp/SearchRequest.csv?jqlQuery=project%3DACSVS+AND+affectedVersion%3Di.mxRT2660+ORDER+BY+priority+DESC%2C+updated+DESC"
# 	# jira_username = "你的Jira用户名"  # 替换为登录NXP Jira的账号
# 	# jira_password = "你的Jira密码"    # 替换为登录密码
# 	# save_filename = "jira_ACSVS_RT2660_issues.csv"  # 保存的文件名

# 	# try:
# 	#     # 发起请求（带Jira基础认证，处理BOM格式CSV）
# 	#     response = requests.get(
# 	#         url,
# 	#         auth=HTTPBasicAuth(jira_username, jira_password),
# 	#         stream=True,
# 	#         headers={"Accept": "text/csv"}  # 明确指定接收CSV格式
# 	#     )
# 	#     response.raise_for_status()  # 若状态码非200，抛出异常

# 	#     # 以二进制模式写入（保留BOM，避免中文乱码）
# 	#     with open(save_filename, "wb") as f:
# 	#         for chunk in response.iter_content(chunk_size=4096):
# 	#             if chunk:
# 	#                 f.write(chunk)

# 	#     print(f"下载成功！文件已保存为：{save_filename}")
# 	# except requests.exceptions.AuthenticationError:
# 	#     print("错误：用户名或密码错误，请检查Jira账号信息")
# 	# except requests.exceptions.RequestException as e:
# 	#     print(f"下载失败：{str(e)}")














# 	# # urllib.request.urlretrieve( url , original_csv)
# 	# response = requests.get(url, stream= True)
# 	# with open(original_csv, 'wb') as f:
# 	# 	for chunk in response.iter_content(chunk_size=1024):
# 	# 		if chunk:
# 	# 			f.write(chunk)




# # if __name__ == '__main__':
# 	# download_excel()
# 	# import requests
# 	# from requests.auth import HTTPBasicAuth

# 	# # Jira登录凭据
# 	# JIRA_USERNAME = "nxa28190"  # 替换为您的用户名
# 	# JIRA_PASSWORD = "Cs@19892000"  # 替换为您的密码
# 	# # 或者使用API令牌（如果Jira配置了双重验证，推荐使用令牌）
# 	# # JIRA_PASSWORD = "your_api_token"

# 	# # CSV文件URL
# 	# url = "https://jira.sw.nxp.com/sr/jira.issueviews:searchrequest-csv-with-bom-all-fields/temp/SearchRequest.csv?jqlQuery=project+%3D+ACSVS+AND+affectedVersion+%3D+i.mxRT2660+ORDER+BY+priority+DESC%2C+updated+DESC"

# 	# # 创建会话并设置认证
# 	# session = requests.Session()

# 	# # 尝试使用HTTP基本认证
# 	# try:
# 	#     response = session.get(
# 	#         url,
# 	#         auth=HTTPBasicAuth(JIRA_USERNAME, JIRA_PASSWORD),
# 	#         headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
# 	#     )
	    
# 	#     # 检查请求是否成功
# 	#     if response.status_code == 200:
# 	#         # 保存CSV文件
# 	#         with open('jira_issues.csv', 'wb') as f:
# 	#             f.write(response.content)
# 	#         print("CSV文件下载成功！")
	        
# 	#         # 显示文件基本信息
# 	#         print(f"文件大小: {len(response.content)} 字节")
# 	#         print(f"保存为: jira_issues.csv")
	        
# 	#     elif response.status_code == 401:
# 	#         print("认证失败：用户名或密码不正确")
# 	#         print("提示：如果启用了双重验证，请使用API令牌而不是密码")
# 	#     elif response.status_code == 403:
# 	#         print("访问被拒绝：您没有权限访问此资源")
# 	#     elif response.status_code == 404:
# 	#         print("文件未找到：URL可能已失效")
# 	#     else:
# 	#         print(f"下载失败，状态码：{response.status_code}")
# 	#         print(f"响应内容：{response.text[:200]}")

# 	# except requests.exceptions.RequestException as e:
# 	#     print(f"网络请求错误：{e}")




# 	# import requests
# 	# from requests.auth import HTTPBasicAuth
# 	# import tkinter as tk
# 	# from tkinter import simpledialog, messagebox
# 	# import getpass

# 	# # CSV文件URL
# 	# url = "https://jira.sw.nxp.com/sr/jira.issueviews:searchrequest-csv-with-bom-current-fields/temp/SearchRequest.csv?jqlQuery=project+%3D+ACSVS+AND+affectedVersion+%3D+i.mxRT2660+ORDER+BY+component+ASC%2C+priority+DESC%2C+updated+DESC"

# 	# def get_credentials_gui():
# 	#     """使用GUI弹窗获取用户名和密码"""
# 	#     root = tk.Tk()
# 	#     root.withdraw()  # 隐藏主窗口
	    
# 	#     # 创建自定义对话框
# 	#     class CredentialsDialog(simpledialog.Dialog):
# 	#         def __init__(self, parent, title=None):
# 	#             self.username = None
# 	#             self.password = None
# 	#             super().__init__(parent, title)
	        
# 	#         def body(self, master):
# 	#             tk.Label(master, text="用户名:").grid(row=0, sticky="w")
# 	#             tk.Label(master, text="密码:").grid(row=1, sticky="w")
	            
# 	#             self.e1 = tk.Entry(master, width=30)
# 	#             self.e2 = tk.Entry(master, width=30, show="*")
	            
# 	#             self.e1.grid(row=0, column=1, padx=5, pady=5)
# 	#             self.e2.grid(row=1, column=1, padx=5, pady=5)
	            
# 	#             return self.e1  # 初始焦点
	        
# 	#         def apply(self):
# 	#             self.username = self.e1.get()
# 	#             self.password = self.e2.get()
	    
# 	#     # 显示对话框
# 	#     dialog = CredentialsDialog(root, "Jira登录凭据")
	    
# 	#     # 销毁根窗口
# 	#     root.destroy()
	    
# 	#     return dialog.username, dialog.password

# 	# def get_credentials_console():
# 	#     """使用控制台获取用户名和密码"""
# 	#     username = input("请输入Jira用户名: ")
# 	#     password = getpass.getpass("请输入Jira密码: ")
# 	#     return username, password

# 	# # 尝试使用GUI获取凭据，如果失败则使用控制台
# 	# try:
# 	#     JIRA_USERNAME, JIRA_PASSWORD = get_credentials_gui()
	    
# 	#     # 检查是否用户取消了对话框
# 	#     if not JIRA_USERNAME or not JIRA_PASSWORD:
# 	#         print("使用控制台输入...")
# 	#         JIRA_USERNAME, JIRA_PASSWORD = get_credentials_console()
	        
# 	# except Exception as e:
# 	#     print(f"GUI不可用，使用控制台输入: {e}")
# 	#     JIRA_USERNAME, JIRA_PASSWORD = get_credentials_console()

# 	# # 确保获取到了凭据
# 	# if not JIRA_USERNAME or not JIRA_PASSWORD:
# 	#     print("错误：未提供用户名或密码")
# 	#     exit(1)

# 	# # 创建会话并设置认证
# 	# session = requests.Session()

# 	# # 尝试使用HTTP基本认证
# 	# try:
# 	#     print("正在连接Jira服务器...")
# 	#     response = session.get(
# 	#         url,
# 	#         auth=HTTPBasicAuth(JIRA_USERNAME, JIRA_PASSWORD),
# 	#         headers={
# 	#             'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
# 	#             'Accept': 'text/csv, */*'
# 	#         },
# 	#         timeout=30
# 	#     )
	    
# 	#     # 检查请求是否成功
# 	#     if response.status_code == 200:
# 	#         # 保存CSV文件
# 	#         filename = 'jira_issues.csv'
# 	#         with open(filename, 'wb') as f:
# 	#             f.write(response.content)
# 	#         print("CSV文件下载成功！")
	        
# 	#         # 显示文件基本信息
# 	#         print(f"文件大小: {len(response.content)} 字节")
# 	#         print(f"保存为: {filename}")
	        
# 	#         # 显示前几行内容预览
# 	#         try:
# 	#             content_preview = response.content.decode('utf-8').split('\n')[:5]
# 	#             print("\n文件前5行预览:")
# 	#             for i, line in enumerate(content_preview):
# 	#                 print(f"{i+1}: {line.strip()}")
# 	#         except UnicodeDecodeError:
# 	#             print("注意：文件内容无法用UTF-8解码")
	        
# 	#     elif response.status_code == 401:
# 	#         print("认证失败：用户名或密码不正确")
# 	#         print("提示：如果启用了双重验证，请使用API令牌而不是密码")
# 	#     elif response.status_code == 403:
# 	#         print("访问被拒绝：您没有权限访问此资源")
# 	#         print("请检查：")
# 	#         print("1. 您是否有权访问ACSVS项目")
# 	#         print("2. 您的账户是否已激活")
# 	#     elif response.status_code == 404:
# 	#         print("文件未找到：URL可能已失效或JQL查询有误")
# 	#         print(f"请验证URL: {url}")
# 	#     else:
# 	#         print(f"下载失败，状态码：{response.status_code}")
# 	#         if response.text:
# 	#             print(f"响应内容：{response.text[:500]}")

# 	# except requests.exceptions.Timeout:
# 	#     print("连接超时：请检查网络连接或尝试使用VPN")
# 	# except requests.exceptions.ConnectionError:
# 	#     print("连接错误：无法连接到服务器，请检查网络连接")
# 	#     print("提示：公司内部资源可能需要VPN连接")
# 	# except requests.exceptions.RequestException as e:
# 	#     print(f"网络请求错误：{e}")
# 	# finally:
# 	#     # 安全清理敏感数据
# 	#     JIRA_USERNAME = None
# 	#     JIRA_PASSWORD = None


# # import pandas as pd
# # import matplotlib.pyplot as plt
# # from collections import Counter
# # import ast
# # import warnings
# # warnings.filterwarnings('ignore')

# # def analyze_jira_components(csv_file_path, output_excel_path):
# #     """
# #     分析Jira CSV文件中的Component数量并生成带图表的Excel报告
    
# #     参数:
# #     csv_file_path: 输入的CSV文件路径
# #     output_excel_path: 输出的Excel文件路径
# #     """
    
# #     try:
# #         # 读取CSV文件
# #         print(f"正在读取文件: {csv_file_path}")
# #         df = pd.read_csv(csv_file_path, encoding='utf-8-sig')
# #         print(f"成功读取数据，共 {len(df)} 行，{len(df.columns)} 列")
        
# #         # 显示所有列名，帮助识别Component列
# #         print("\n数据列名:")
# #         for i, col in enumerate(df.columns):
# #             print(f"{i+1}: {col}")
        
# #         # 尝试自动识别Component列
# #         component_col = None
# #         possible_names = ['Component', 'Components', 'component', 'components', 
# #                          'COMPONENT', 'COMPONENTS', '模块', '组件']
        
# #         for col in df.columns:
# #             if col in possible_names:
# #                 component_col = col
# #                 break
                
# #         # 如果没有自动识别到，让用户选择
# #         if component_col is None:
# #             print("\n未自动识别到Component列，请从以下列中选择:")
# #             for i, col in enumerate(df.columns):
# #                 print(f"{i+1}: {col}")
            
# #             try:
# #                 choice = int(input("请输入列号: ")) - 1
# #                 if 0 <= choice < len(df.columns):
# #                     component_col = df.columns[choice]
# #                 else:
# #                     print("无效选择，将尝试查找包含'component'的列名")
# #                     for col in df.columns:
# #                         if 'component' in col.lower():
# #                             component_col = col
# #                             break
# #             except ValueError:
# #                 print("输入无效，将尝试查找包含'component'的列名")
# #                 for col in df.columns:
# #                     if 'component' in col.lower():
# #                         component_col = col
# #                         break
        
# #         if component_col is None:
# #             print("错误：无法确定Component列，请检查CSV文件结构")
# #             return False
            
# #         print(f"\n使用列 '{component_col}' 进行Component统计")
        
# #         # 统计Component数量
# #         component_stats = analyze_components(df, component_col)
        
# #         if not component_stats:
# #             print("未找到有效的Component数据")
# #             return False
        
# #         # 创建Excel文件
# #         create_excel_report(df, component_stats, component_col, output_excel_path)
        
# #         print(f"\n分析完成！结果已保存到: {output_excel_path}")
# #         return True
        
# #     except Exception as e:
# #         print(f"分析过程中出错: {e}")
# #         return False

# # def analyze_components(df, component_col):
# #     """
# #     分析Component数据并返回统计结果
# #     """
# #     all_components = []
    
# #     # 处理Component列（可能是字符串、列表或NaN）
# #     for components in df[component_col].dropna():
# #         if isinstance(components, str):
# #             # 尝试解析字符串（可能是列表形式的字符串）
# #             try:
# #                 # 如果是类似 "['Comp1', 'Comp2']" 的格式
# #                 if components.startswith('[') and components.endswith(']'):
# #                     comp_list = ast.literal_eval(components)
# #                     if isinstance(comp_list, list):
# #                         all_components.extend(comp_list)
# #                     else:
# #                         all_components.append(str(comp_list))
# #                 # 如果是分号分隔的
# #                 elif ';' in components:
# #                     all_components.extend([comp.strip() for comp in components.split(';') if comp.strip()])
# #                 # 如果是逗号分隔的
# #                 elif ',' in components:
# #                     all_components.extend([comp.strip() for comp in components.split(',') if comp.strip()])
# #                 else:
# #                     all_components.append(components.strip())
# #             except:
# #                 # 如果解析失败，直接作为单个组件处理
# #                 all_components.append(components.strip())
# #         elif isinstance(components, list):
# #             all_components.extend(components)
# #         else:
# #             all_components.append(str(components))
    
# #     # 过滤空字符串
# #     all_components = [comp for comp in all_components if comp and comp.strip()]
    
# #     if not all_components:
# #         return None
    
# #     # 统计数量
# #     component_counts = Counter(all_components)
    
# #     # 转换为DataFrame并排序
# #     stats_df = pd.DataFrame({
# #         'Component': list(component_counts.keys()),
# #         'Count': list(component_counts.values())
# #     }).sort_values('Count', ascending=False)
    
# #     print(f"\n找到 {len(stats_df)} 个不同的Component")
# #     print("\nComponent统计前10名:")
# #     print(stats_df.head(10).to_string(index=False))
    
# #     return stats_df

# # def create_excel_report(df, component_stats, component_col, output_path):
# #     """
# #     创建包含统计数据和图表的Excel报告
# #     """
# #     # 创建Excel写入器
# #     with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
# #         # 1. 写入原始数据
# #         df.to_excel(writer, sheet_name='原始数据', index=False)
        
# #         # 2. 写入Component统计
# #         component_stats.to_excel(writer, sheet_name='Component统计', index=False)
        
# #         # 3. 创建汇总统计表
# #         summary_data = {
# #             '统计项': ['总Issue数', '有Component的Issue数', '无Component的Issue数', 
# #                      '唯一Component数', '最多Component的Issue'],
# #             '数值': [
# #                 len(df),
# #                 df[component_col].notna().sum(),
# #                 df[component_col].isna().sum(),
# #                 len(component_stats),
# #                 f"{component_stats.iloc[0]['Component']} ({component_stats.iloc[0]['Count']}次)"
# #             ]
# #         }
# #         summary_df = pd.DataFrame(summary_data)
# #         summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
        
# #         # 获取工作簿和工作表以添加图表
# #         workbook = writer.book
# #         stats_sheet = writer.sheets['Component统计']
        
# #         # 创建图表工作表
# #         chart_sheet = workbook.create_sheet(title='图表')
        
# #         # 生成图表
# #         create_charts(component_stats, chart_sheet, workbook)
        
# #         # 调整列宽
# #         for sheet_name in writer.sheets:
# #             worksheet = writer.sheets[sheet_name]
# #             for column in worksheet.columns:
# #                 max_length = 0
# #                 column_letter = column[0].column_letter
# #                 for cell in column:
# #                     try:
# #                         if len(str(cell.value)) > max_length:
# #                             max_length = len(str(cell.value))
# #                     except:
# #                         pass
# #                 adjusted_width = min(max_length + 2, 50)
# #                 worksheet.column_dimensions[column_letter].width = adjusted_width

# # def create_charts(component_stats, chart_sheet, workbook):
# #     """
# #     在Excel中创建图表
# #     """
# #     from openpyxl.chart import BarChart, PieChart, Reference
    
# #     # 限制显示的Component数量（避免图表过于拥挤）
# #     max_display = min(15, len(component_stats))
# #     display_data = component_stats.head(max_display)
    
# #     # 1. 创建条形图
# #     bar_chart = BarChart()
# #     bar_chart.title = f"Top {max_display} Components 统计"
# #     bar_chart.style = 10
# #     bar_chart.y_axis.title = '数量'
# #     bar_chart.x_axis.title = 'Component'
    
# #     # 数据引用
# #     data = Reference(chart_sheet, min_col=2, min_row=1, max_row=max_display+1)
# #     categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=max_display+1)
    
# #     bar_chart.add_data(data, titles_from_data=True)
# #     bar_chart.set_categories(categories)
    
# #     # 将图表添加到工作表
# #     chart_sheet.add_chart(bar_chart, "A1")
    
# #     # 2. 创建饼图（显示前8个）
# #     pie_max = min(8, len(component_stats))
# #     pie_data = component_stats.head(pie_max)
    
# #     # 将数据写入工作表用于饼图
# #     for i, (_, row) in enumerate(pie_data.iterrows(), 1):
# #         chart_sheet.cell(row=i+20, column=1, value=row['Component'])
# #         chart_sheet.cell(row=i+20, column=2, value=row['Count'])
    
# #     pie_chart = PieChart()
# #     pie_chart.title = f"Top {pie_max} Components 分布"
    
# #     pie_data_ref = Reference(chart_sheet, min_col=2, min_row=21, max_row=20+pie_max)
# #     pie_labels_ref = Reference(chart_sheet, min_col=1, min_row=21, max_row=20+pie_max)
    
# #     pie_chart.add_data(pie_data_ref, titles_from_data=True)
# #     pie_chart.set_categories(pie_labels_ref)
    
# #     chart_sheet.add_chart(pie_chart, "A20")

# # def main():
# #     """
# #     主函数 - 执行完整的分析流程
# #     """
# #     print("Jira Component 统计分析工具")
# #     print("=" * 50)
    
# #     # 输入文件路径
# #     csv_file = input("请输入CSV文件路径 (直接回车使用 'jira_issues.csv'): ").strip()
# #     if not csv_file:
# #         csv_file = 'jira_issues.csv'
    
# #     # 输出文件路径
# #     output_file = input("请输入输出Excel文件路径 (直接回车使用 'jira_analysis.xlsx'): ").strip()
# #     if not output_file:
# #         output_file = 'jira_analysis.xlsx'
    
# #     # 执行分析
# #     success = analyze_jira_components(csv_file, output_file)
    
# #     if success:
# #         print(f"\n✅ 分析完成！")
# #         print(f"📊 生成的Excel文件包含:")
# #         print(f"   - 原始数据表")
# #         print(f"   - Component统计表") 
# #         print(f"   - 汇总统计表")
# #         print(f"   - 条形图和饼图")
# #     else:
# #         print(f"\n❌ 分析失败，请检查输入文件和数据格式")

# # if __name__ == "__main__":
# #     # 检查必要的库
# #     try:
# #         import openpyxl
# #     except ImportError:
# #         print("缺少必要的库，正在安装...")
# #         import subprocess
# #         import sys
# #         subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "matplotlib"])
# #         print("库安装完成，请重新运行脚本")
# #         exit()
    
# #     main()



# # import pandas as pd
# # import matplotlib.pyplot as plt
# # from collections import Counter
# # import ast
# # import warnings
# # import os
# # from datetime import datetime

# # warnings.filterwarnings('ignore')

# # def analyze_jira_components():
# #     """
# #     分析Jira CSV文件中的Component数量并生成带图表的Excel报告
# #     """
    
# #     # 使用默认CSV文件名
# #     csv_file_path = 'jira_issues.csv'
    
# #     # 生成带时间戳的输出文件名
# #     timestamp = datetime.now().strftime("%Y%m%d_%H%M")
# #     output_excel_path = f'jira_analysis_{timestamp}.xlsx'
    
# #     try:
# #         # 检查CSV文件是否存在
# #         if not os.path.exists(csv_file_path):
# #             print(f"错误：找不到CSV文件 '{csv_file_path}'")
# #             print("请确保CSV文件与脚本在同一目录下")
# #             return False
        
# #         # 读取CSV文件
# #         print(f"正在读取文件: {csv_file_path}")
# #         df = pd.read_csv(csv_file_path, encoding='utf-8-sig')
# #         print(f"成功读取数据，共 {len(df)} 行，{len(df.columns)} 列")
        
# #         # 显示所有列名，帮助识别Component列
# #         print("\n数据列名:")
# #         for i, col in enumerate(df.columns):
# #             print(f"{i+1}: {col}")
        
# #         # 尝试自动识别Component列
# #         component_col = None
# #         possible_names = ['Component', 'Components', 'component', 'components', 
# #                          'COMPONENT', 'COMPONENTS', '模块', '组件']
        
# #         for col in df.columns:
# #             if col in possible_names:
# #                 component_col = col
# #                 break
                
# #         # 如果没有自动识别到，尝试查找包含'component'的列名
# #         if component_col is None:
# #             for col in df.columns:
# #                 if 'component' in col.lower():
# #                     component_col = col
# #                     break
        
# #         if component_col is None:
# #             print("错误：无法确定Component列，请检查CSV文件结构")
# #             return False
            
# #         print(f"\n使用列 '{component_col}' 进行Component统计")
        
# #         # 统计Component数量
# #         component_stats = analyze_components(df, component_col)
        
# #         if not component_stats:
# #             print("未找到有效的Component数据")
# #             return False
        
# #         # 创建Excel文件
# #         create_excel_report(df, component_stats, component_col, output_excel_path)
        
# #         print(f"\n分析完成！结果已保存到: {output_excel_path}")
# #         return True
        
# #     except Exception as e:
# #         print(f"分析过程中出错: {e}")
# #         return False

# # def analyze_components(df, component_col):
# #     """
# #     分析Component数据并返回统计结果
# #     """
# #     all_components = []
    
# #     # 处理Component列（可能是字符串、列表或NaN）
# #     for components in df[component_col].dropna():
# #         if isinstance(components, str):
# #             # 尝试解析字符串（可能是列表形式的字符串）
# #             try:
# #                 # 如果是类似 "['Comp1', 'Comp2']" 的格式
# #                 if components.startswith('[') and components.endswith(']'):
# #                     comp_list = ast.literal_eval(components)
# #                     if isinstance(comp_list, list):
# #                         all_components.extend(comp_list)
# #                     else:
# #                         all_components.append(str(comp_list))
# #                 # 如果是分号分隔的
# #                 elif ';' in components:
# #                     all_components.extend([comp.strip() for comp in components.split(';') if comp.strip()])
# #                 # 如果是逗号分隔的
# #                 elif ',' in components:
# #                     all_components.extend([comp.strip() for comp in components.split(',') if comp.strip()])
# #                 else:
# #                     all_components.append(components.strip())
# #             except:
# #                 # 如果解析失败，直接作为单个组件处理
# #                 all_components.append(components.strip())
# #         elif isinstance(components, list):
# #             all_components.extend(components)
# #         else:
# #             all_components.append(str(components))
    
# #     # 过滤空字符串
# #     all_components = [comp for comp in all_components if comp and comp.strip()]
    
# #     if not all_components:
# #         return None
    
# #     # 统计数量
# #     component_counts = Counter(all_components)
    
# #     # 转换为DataFrame并排序
# #     stats_df = pd.DataFrame({
# #         'Component': list(component_counts.keys()),
# #         'Count': list(component_counts.values())
# #     }).sort_values('Count', ascending=False)
    
# #     print(f"\n找到 {len(stats_df)} 个不同的Component")
# #     print("\nComponent统计前10名:")
# #     print(stats_df.head(10).to_string(index=False))
    
# #     return stats_df

# # def create_excel_report(df, component_stats, component_col, output_path):
# #     """
# #     创建包含统计数据和图表的Excel报告
# #     """
# #     # 创建Excel写入器
# #     with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
# #         # 1. 写入原始数据
# #         df.to_excel(writer, sheet_name='原始数据', index=False)
        
# #         # 2. 写入Component统计
# #         component_stats.to_excel(writer, sheet_name='Component统计', index=False)
        
# #         # 3. 创建汇总统计表
# #         summary_data = {
# #             '统计项': ['总Issue数', '有Component的Issue数', '无Component的Issue数', 
# #                      '唯一Component数', '最多Component的Issue', '分析时间'],
# #             '数值': [
# #                 len(df),
# #                 df[component_col].notna().sum(),
# #                 df[component_col].isna().sum(),
# #                 len(component_stats),
# #                 f"{component_stats.iloc[0]['Component']} ({component_stats.iloc[0]['Count']}次)",
# #                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")
# #             ]
# #         }
# #         summary_df = pd.DataFrame(summary_data)
# #         summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
        
# #         # 获取工作簿和工作表以添加图表
# #         workbook = writer.book
# #         stats_sheet = writer.sheets['Component统计']
        
# #         # 创建图表工作表
# #         chart_sheet = workbook.create_sheet(title='图表')
        
# #         # 生成图表
# #         create_charts(component_stats, chart_sheet, workbook)
        
# #         # 调整列宽
# #         for sheet_name in writer.sheets:
# #             worksheet = writer.sheets[sheet_name]
# #             for column in worksheet.columns:
# #                 max_length = 0
# #                 column_letter = column[0].column_letter
# #                 for cell in column:
# #                     try:
# #                         if len(str(cell.value)) > max_length:
# #                             max_length = len(str(cell.value))
# #                     except:
# #                         pass
# #                 adjusted_width = min(max_length + 2, 50)
# #                 worksheet.column_dimensions[column_letter].width = adjusted_width

# # def create_charts(component_stats, chart_sheet, workbook):
# #     """
# #     在Excel中创建图表
# #     """
# #     from openpyxl.chart import BarChart, PieChart, Reference
    
# #     # 限制显示的Component数量（避免图表过于拥挤）
# #     max_display = min(15, len(component_stats))
# #     display_data = component_stats.head(max_display)
    
# #     # 将数据写入工作表用于图表
# #     # 写入表头
# #     chart_sheet['A1'] = 'Component'
# #     chart_sheet['B1'] = 'Count'
    
# #     # 写入数据
# #     for i, (_, row) in enumerate(display_data.iterrows(), 2):
# #         chart_sheet[f'A{i}'] = row['Component']
# #         chart_sheet[f'B{i}'] = row['Count']
    
# #     # 1. 创建条形图
# #     bar_chart = BarChart()
# #     bar_chart.title = f"Top {max_display} Components 统计"
# #     bar_chart.style = 10
# #     bar_chart.y_axis.title = '数量'
# #     bar_chart.x_axis.title = 'Component'
    
# #     # 数据引用
# #     data = Reference(chart_sheet, min_col=2, min_row=1, max_row=max_display+1)
# #     categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=max_display+1)
    
# #     bar_chart.add_data(data, titles_from_data=True)
# #     bar_chart.set_categories(categories)
    
# #     # 将条形图添加到工作表
# #     chart_sheet.add_chart(bar_chart, "D1")
    
# #     # 2. 创建饼图（显示前8个）
# #     pie_max = min(8, len(component_stats))
# #     pie_data = component_stats.head(pie_max)
    
# #     # 将数据写入工作表用于饼图
# #     chart_sheet['A20'] = 'Component'
# #     chart_sheet['B20'] = 'Count'
    
# #     for i, (_, row) in enumerate(pie_data.iterrows(), 1):
# #         chart_sheet.cell(row=i+20, column=1, value=row['Component'])
# #         chart_sheet.cell(row=i+20, column=2, value=row['Count'])
    
# #     pie_chart = PieChart()
# #     pie_chart.title = f"Top {pie_max} Components 分布"
    
# #     pie_data_ref = Reference(chart_sheet, min_col=2, min_row=21, max_row=20+pie_max)
# #     pie_labels_ref = Reference(chart_sheet, min_col=1, min_row=21, max_row=20+pie_max)
    
# #     pie_chart.add_data(pie_data_ref, titles_from_data=True)
# #     pie_chart.set_categories(pie_labels_ref)
    
# #     chart_sheet.add_chart(pie_chart, "D20")

# # def main():
# #     """
# #     主函数 - 执行完整的分析流程
# #     """
# #     print("Jira Component 统计分析工具")
# #     print("=" * 50)
    
# #     # 执行分析
# #     success = analyze_jira_components()
    
# #     if success:
# #         print(f"\n✅ 分析完成！")
# #         print(f"📊 生成的Excel文件包含:")
# #         print(f"   - 原始数据表")
# #         print(f"   - Component统计表") 
# #         print(f"   - 汇总统计表")
# #         print(f"   - 条形图和饼图")
# #     else:
# #         print(f"\n❌ 分析失败，请检查输入文件和数据格式")

# # if __name__ == "__main__":
# #     # 检查必要的库
# #     try:
# #         import openpyxl
# #     except ImportError:
# #         print("缺少必要的库，正在安装...")
# #         import subprocess
# #         import sys
# #         subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "matplotlib"])
# #         print("库安装完成，请重新运行脚本")
# #         exit()
    
# #     main()


# # # ### 主要改进：

# # # 1. **自动使用默认CSV文件**：
# # #    - 不再提示输入CSV文件路径，直接使用 `jira_issues.csv`
# # #    - 添加了文件存在性检查

# # # 2. **自动生成带时间戳的输出文件名**：
# # #    - 使用当前日期和时间生成文件名，格式为 `jira_analysis_YYYYMMDD_HHMM.xlsx`
# # #    - 例如：`jira_analysis_20231215_1430.xlsx`

# # # 3. **在汇总统计中添加分析时间**：
# # #    - 在汇总统计表中记录分析执行的具体时间

# # # 4. **优化图表布局**：
# # #    - 调整了条形图和饼图在工作表中的位置，避免重叠

# # # ### 使用说明：

# # # 1. 确保 `jira_issues.csv` 文件与脚本在同一目录下
# # # 2. 直接运行脚本，无需任何交互输入：
# # #    ```bash
# # #    python jira_analyzer.py
# # #    ```
# # # 3. 脚本会自动生成带时间戳的Excel文件，包含完整的分析结果和图表

# # # ### 输出文件命名规则：

# # # - 格式：`jira_analysis_年月日_时分.xlsx`
# # # - 示例：`jira_analysis_20231215_1430.xlsx`

# # # 这样每次运行脚本都会生成一个唯一的输出文件，方便您跟踪不同时间的分析结果。



# import requests
# from requests.auth import HTTPBasicAuth
# import tkinter as tk
# from tkinter import simpledialog, messagebox
# import pandas as pd
# import matplotlib.pyplot as plt
# from collections import Counter
# import ast
# import warnings
# import os
# from datetime import datetime
# import getpass

# warnings.filterwarnings('ignore')

# def get_credentials():
#     """使用GUI弹窗获取用户名和密码"""
#     try:
#         root = tk.Tk()
#         root.withdraw()  # 隐藏主窗口
        
#         # 创建自定义对话框
#         class CredentialsDialog(simpledialog.Dialog):
#             def __init__(self, parent, title=None):
#                 self.username = None
#                 self.password = None
#                 super().__init__(parent, title)
            
#             def body(self, master):
#                 tk.Label(master, text="Jira用户名:").grid(row=0, sticky="w")
#                 tk.Label(master, text="Jira密码:").grid(row=1, sticky="w")
                
#                 self.e1 = tk.Entry(master, width=30)
#                 self.e2 = tk.Entry(master, width=30, show="*")
                
#                 self.e1.grid(row=0, column=1, padx=5, pady=5)
#                 self.e2.grid(row=1, column=1, padx=5, pady=5)
                
#                 return self.e1  # 初始焦点
            
#             def apply(self):
#                 self.username = self.e1.get()
#                 self.password = self.e2.get()
        
#         # 显示对话框
#         dialog = CredentialsDialog(root, "Jira登录凭据")
        
#         # 销毁根窗口
#         root.destroy()
        
#         return dialog.username, dialog.password
        
#     except Exception as e:
#         print(f"GUI不可用，使用控制台输入: {e}")
#         username = input("请输入Jira用户名: ")
#         password = getpass.getpass("请输入Jira密码: ")
#         return username, password

# def download_jira_csv():
#     """下载Jira CSV文件"""
#     # Jira CSV文件URL
#     url = "https://jira.sw.nxp.com/sr/jira.issueviews:searchrequest-csv-with-bom-all-fields/temp/SearchRequest.csv?jqlQuery=project+%3D+ACSVS+AND+affectedVersion+%3D+i.mxRT2660+ORDER+BY+priority+DESC%2C+updated+DESC"
    
#     # 获取用户名和密码
#     username, password = get_credentials()
    
#     if not username or not password:
#         print("错误：未提供用户名或密码")
#         return None
    
#     # 创建会话并设置认证
#     session = requests.Session()
    
#     # 生成带时间戳的文件名
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M")
#     csv_filename = f'jira_issues_{timestamp}.csv'
    
#     try:
#         print("正在连接Jira服务器...")
#         response = session.get(
#             url,
#             auth=HTTPBasicAuth(username, password),
#             headers={
#                 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
#                 'Accept': 'text/csv, */*'
#             },
#             timeout=30
#         )
        
#         # 检查请求是否成功
#         if response.status_code == 200:
#             # 保存CSV文件
#             with open(csv_filename, 'wb') as f:
#                 f.write(response.content)
#             print(f"CSV文件下载成功！保存为: {csv_filename}")
            
#             # 显示文件基本信息
#             print(f"文件大小: {len(response.content)} 字节")
            
#             # 显示前几行内容预览
#             try:
#                 content_preview = response.content.decode('utf-8-sig').split('\n')[:5]
#                 print("\n文件前5行预览:")
#                 for i, line in enumerate(content_preview):
#                     print(f"{i+1}: {line.strip()}")
#             except UnicodeDecodeError:
#                 print("注意：文件内容无法用UTF-8解码")
            
#             return csv_filename
            
#         elif response.status_code == 401:
#             print("认证失败：用户名或密码不正确")
#             print("提示：如果启用了双重验证，请使用API令牌而不是密码")
#         elif response.status_code == 403:
#             print("访问被拒绝：您没有权限访问此资源")
#         elif response.status_code == 404:
#             print("文件未找到：URL可能已失效或JQL查询有误")
#         else:
#             print(f"下载失败，状态码：{response.status_code}")
#             if response.text:
#                 print(f"响应内容：{response.text[:500]}")
        
#         return None
        
#     except requests.exceptions.Timeout:
#         print("连接超时：请检查网络连接或尝试使用VPN")
#     except requests.exceptions.ConnectionError:
#         print("连接错误：无法连接到服务器，请检查网络连接")
#         print("提示：公司内部资源可能需要VPN连接")
#     except requests.exceptions.RequestException as e:
#         print(f"网络请求错误：{e}")
    
#     return None

# def analyze_jira_components(csv_file_path):
#     """
#     分析Jira CSV文件中的Component数量并生成带图表的Excel报告
#     """
#     # 生成带时间戳的输出文件名
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M")
#     output_excel_path = f'jira_analysis_{timestamp}.xlsx'
    
#     try:
#         # 检查CSV文件是否存在
#         if not os.path.exists(csv_file_path):
#             print(f"错误：找不到CSV文件 '{csv_file_path}'")
#             return False
        
#         # 读取CSV文件
#         print(f"正在读取文件: {csv_file_path}")
#         df = pd.read_csv(csv_file_path, encoding='utf-8-sig')
#         print(f"成功读取数据，共 {len(df)} 行，{len(df.columns)} 列")
        
#         # 显示所有列名，帮助识别Component列
#         print("\n数据列名:")
#         for i, col in enumerate(df.columns):
#             print(f"{i+1}: {col}")
        
#         # 尝试自动识别Component列
#         component_col = None
#         possible_names = ['Component', 'Components', 'component', 'components', 
#                          'COMPONENT', 'COMPONENTS', '模块', '组件']
        
#         for col in df.columns:
#             if col in possible_names:
#                 component_col = col
#                 break
                
#         # 如果没有自动识别到，尝试查找包含'component'的列名
#         if component_col is None:
#             for col in df.columns:
#                 if 'component' in col.lower():
#                     component_col = col
#                     break
        
#         # 如果仍然没有找到，让用户选择
#         if component_col is None:
#             print("\n未自动识别到Component列，请从以下列中选择:")
#             for i, col in enumerate(df.columns):
#                 print(f"{i+1}: {col}")
            
#             try:
#                 choice = int(input("请输入列号: ")) - 1
#                 if 0 <= choice < len(df.columns):
#                     component_col = df.columns[choice]
#                 else:
#                     print("无效选择，将使用第一列")
#                     component_col = df.columns[0]
#             except (ValueError, IndexError):
#                 print("输入无效，将使用第一列")
#                 component_col = df.columns[0]
            
#         print(f"\n使用列 '{component_col}' 进行Component统计")
        
#         # 统计Component数量
#         component_stats = analyze_components(df, component_col)
        
#         # 检查component_stats是否为空DataFrame
#         if component_stats is None or (hasattr(component_stats, 'empty') and component_stats.empty):
#             print("未找到有效的Component数据")
#             return False
        
#         # 创建Excel文件
#         create_excel_report(df, component_stats, component_col, output_excel_path)
        
#         print(f"\n分析完成！结果已保存到: {output_excel_path}")
#         return output_excel_path
        
#     except Exception as e:
#         print(f"分析过程中出错: {e}")
#         import traceback
#         traceback.print_exc()
#         return False

# def analyze_components(df, component_col):
#     """
#     分析Component数据并返回统计结果
#     """
#     all_components = []
    
#     # 检查列是否存在
#     if component_col not in df.columns:
#         print(f"错误：列 '{component_col}' 不存在于数据中")
#         return None
    
#     # 处理Component列（可能是字符串、列表或NaN）
#     for components in df[component_col].dropna():
#         if isinstance(components, str):
#             # 尝试解析字符串（可能是列表形式的字符串）
#             try:
#                 # 如果是类似 "['Comp1', 'Comp2']" 的格式
#                 if components.startswith('[') and components.endswith(']'):
#                     comp_list = ast.literal_eval(components)
#                     if isinstance(comp_list, list):
#                         all_components.extend(comp_list)
#                     else:
#                         all_components.append(str(comp_list))
#                 # 如果是分号分隔的
#                 elif ';' in components:
#                     all_components.extend([comp.strip() for comp in components.split(';') if comp.strip()])
#                 # 如果是逗号分隔的
#                 elif ',' in components:
#                     all_components.extend([comp.strip() for comp in components.split(',') if comp.strip()])
#                 else:
#                     all_components.append(components.strip())
#             except:
#                 # 如果解析失败，直接作为单个组件处理
#                 all_components.append(components.strip())
#         elif isinstance(components, list):
#             all_components.extend(components)
#         else:
#             all_components.append(str(components))
    
#     # 过滤空字符串
#     all_components = [comp for comp in all_components if comp and comp.strip()]
    
#     if not all_components:
#         print("没有找到有效的Component数据")
#         return None
    
#     # 统计数量
#     component_counts = Counter(all_components)
    
#     # 转换为DataFrame并排序
#     stats_df = pd.DataFrame({
#         'Component': list(component_counts.keys()),
#         'Count': list(component_counts.values())
#     }).sort_values('Count', ascending=False)
    
#     print(f"\n找到 {len(stats_df)} 个不同的Component")
#     print("\nComponent统计前10名:")
#     print(stats_df.head(10).to_string(index=False))
    
#     return stats_df

# def create_excel_report(df, component_stats, component_col, output_path):
#     """
#     创建包含统计数据和图表的Excel报告
#     """
#     # 创建Excel写入器
#     with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
#         # 1. 写入原始数据
#         df.to_excel(writer, sheet_name='原始数据', index=False)
        
#         # 2. 写入Component统计
#         component_stats.to_excel(writer, sheet_name='Component统计', index=False)
        
#         # 3. 创建汇总统计表
#         summary_data = {
#             '统计项': ['总Issue数', '有Component的Issue数', '无Component的Issue数', 
#                      '唯一Component数', '最多Component的Issue', '分析时间'],
#             '数值': [
#                 len(df),
#                 df[component_col].notna().sum(),
#                 df[component_col].isna().sum(),
#                 len(component_stats),
#                 f"{component_stats.iloc[0]['Component']} ({component_stats.iloc[0]['Count']}次)",
#                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#             ]
#         }
#         summary_df = pd.DataFrame(summary_data)
#         summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
        
#         # 获取工作簿和工作表以添加图表
#         workbook = writer.book
        
#         # 创建图表工作表
#         chart_sheet = workbook.create_sheet(title='图表')
        
#         # 生成图表
#         create_charts(component_stats, chart_sheet, workbook)
        
#         # 调整列宽
#         for sheet_name in writer.sheets:
#             worksheet = writer.sheets[sheet_name]
#             for column in worksheet.columns:
#                 max_length = 0
#                 column_letter = column[0].column_letter
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = min(max_length + 2, 50)
#                 worksheet.column_dimensions[column_letter].width = adjusted_width

# def create_charts(component_stats, chart_sheet, workbook):
#     """
#     在Excel中创建图表
#     """
#     from openpyxl.chart import BarChart, PieChart, Reference
    
#     # 限制显示的Component数量（避免图表过于拥挤）
#     max_display = min(15, len(component_stats))
#     display_data = component_stats.head(max_display)
    
#     # 将数据写入工作表用于图表
#     # 写入表头
#     chart_sheet['A1'] = 'Component'
#     chart_sheet['B1'] = 'Count'
    
#     # 写入数据
#     for i, (_, row) in enumerate(display_data.iterrows(), 2):
#         chart_sheet[f'A{i}'] = row['Component']
#         chart_sheet[f'B{i}'] = row['Count']
    
#     # 1. 创建条形图
#     bar_chart = BarChart()
#     bar_chart.title = f"Top {max_display} Components 统计"
#     bar_chart.style = 10
#     bar_chart.y_axis.title = '数量'
#     bar_chart.x_axis.title = 'Component'
    
#     # 数据引用
#     data = Reference(chart_sheet, min_col=2, min_row=1, max_row=max_display+1)
#     categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=max_display+1)
    
#     bar_chart.add_data(data, titles_from_data=True)
#     bar_chart.set_categories(categories)
    
#     # 将条形图添加到工作表
#     chart_sheet.add_chart(bar_chart, "D1")
    
#     # 2. 创建饼图（显示前8个）
#     pie_max = min(8, len(component_stats))
#     pie_data = component_stats.head(pie_max)
    
#     # 将数据写入工作表用于饼图
#     chart_sheet['A20'] = 'Component'
#     chart_sheet['B20'] = 'Count'
    
#     for i, (_, row) in enumerate(pie_data.iterrows(), 1):
#         chart_sheet.cell(row=i+20, column=1, value=row['Component'])
#         chart_sheet.cell(row=i+20, column=2, value=row['Count'])
    
#     pie_chart = PieChart()
#     pie_chart.title = f"Top {pie_max} Components 分布"
    
#     pie_data_ref = Reference(chart_sheet, min_col=2, min_row=21, max_row=20+pie_max)
#     pie_labels_ref = Reference(chart_sheet, min_col=1, min_row=21, max_row=20+pie_max)
    
#     pie_chart.add_data(pie_data_ref, titles_from_data=True)
#     pie_chart.set_categories(pie_labels_ref)
    
#     chart_sheet.add_chart(pie_chart, "D20")

# def main():
#     """
#     主函数 - 执行完整的流程
#     """
#     print("Jira CSV下载与分析工具")
#     print("=" * 50)
    
#     # 1. 下载CSV文件
#     print("步骤 1/2: 下载Jira CSV文件")
#     csv_file = download_jira_csv()
    
#     if not csv_file:
#         print("CSV文件下载失败，程序退出")
#         return
    
#     print(f"\n{'='*50}")
    
#     # 2. 分析CSV文件并生成Excel报告
#     print("步骤 2/2: 分析CSV文件并生成Excel报告")
#     excel_file = analyze_jira_components(csv_file)
    
#     if excel_file:
#         print(f"\n✅ 所有任务完成！")
#         print(f"📄 下载的CSV文件: {csv_file}")
#         print(f"📊 生成的Excel文件: {excel_file}")
#         print(f"\nExcel文件包含:")
#         print(f"   - 原始数据表")
#         print(f"   - Component统计表") 
#         print(f"   - 汇总统计表")
#         print(f"   - 条形图和饼图")
#     else:
#         print(f"\n❌ Excel报告生成失败")

# if __name__ == "__main__":
#     # 检查必要的库
#     try:
#         import openpyxl
#     except ImportError:
#         print("缺少必要的库，正在安装...")
#         import subprocess
#         import sys
#         subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "matplotlib", "pandas", "requests"])
#         print("库安装完成，请重新运行脚本")
#         exit()
    
#     main()

# # ```

# # ### 功能特点：

# # 1. **密码弹窗输入**：
# #    - 使用 `tkinter` 创建图形界面输入用户名和密码
# #    - 如果GUI不可用，自动回退到控制台输入

# # 2. **自动下载CSV文件**：
# #    - 使用 `requests` 库下载Jira CSV文件
# #    - 处理HTTP基本认证
# #    - 详细的错误处理和状态反馈

# # 3. **自动分析并生成Excel报告**：
# #    - 自动识别Component列
# #    - 统计Component数量并排序
# #    - 生成包含图表的多工作表Excel文件

# # 4. **带时间戳的文件命名**：
# #    - CSV文件：`jira_issues_YYYYMMDD_HHMM.csv`
# #    - Excel文件：`jira_analysis_YYYYMMDD_HHMM.xlsx`

# # ### 使用说明：

# # 1. 运行脚本：
# #    ```bash
# #    python jira_download_analyze.py
# #    ```

# # 2. 在弹窗中输入Jira用户名和密码：
# #    - 如果启用了双重验证，请使用API令牌而不是密码

# # 3. 脚本会自动：
# #    - 下载CSV文件并保存为带时间戳的文件
# #    - 分析CSV文件中的Component数据
# #    - 生成带图表的Excel报告

# # ### 输出文件：

# # - **CSV文件**：包含从Jira下载的原始数据
# # - **Excel文件**：包含：
# #   - 原始数据表
# #   - Component统计表
# #   - 汇总统计表
# #   - 图表页面（条形图和饼图）

# # ### 依赖库：

# # - `requests`：用于HTTP请求
# # - `pandas`：用于数据处理
# # - `openpyxl`：用于生成Excel文件
# # - `matplotlib`：用于绘图（虽然主要使用openpyxl的图表功能）
# # - `tkinter`：用于GUI输入（通常Python自带）

# # 如果缺少任何库，脚本会自动尝试安装。

# # 这个完整的脚本将帮助您一键完成从Jira下载CSV文件到生成分析报告的全过程，所有文件都会自动使用当前日期时间作为后缀，避免覆盖之前的文件。



# #===================================================================================================================
# #===================================================================================================================
# #===================================================================================================================
# #     backup code 1, richard 
# #===================================================================================================================
# #===================================================================================================================
# #===================================================================================================================

# # 这个错误表明配置文件中缺少 `files` 部分。让我们修复这个问题，确保代码能够正确处理缺失的配置项，并提供默认值：

# # ```python
# import requests
# from requests.auth import HTTPBasicAuth
# import tkinter as tk
# from tkinter import simpledialog, messagebox
# import pandas as pd
# import matplotlib.pyplot as plt
# from collections import Counter
# import ast
# import warnings
# import os
# from datetime import datetime
# import getpass
# import yaml
# import time
# import sys

# warnings.filterwarnings('ignore')

# def load_config(config_file='config.yaml'):
#     """从YAML文件加载配置，并提供默认值"""
#     # 默认配置
#     default_config = {
#         'jira': {
#             'url': 'https://jira.sw.nxp.com/sr/jira.issueviews:searchrequest-csv-with-bom-all-fields/temp/SearchRequest.csv?jqlQuery=project+%3D+ACSVS+AND+affectedVersion+%3D+i.mxRT2660+ORDER+BY+priority+DESC%2C+updated+DESC',
#             'timeout': 60,
#             'retries': 3,
#             'retry_delay': 5
#         },
#         'analysis': {
#             'component_column_names': ['Component', 'Components', 'component', 'components', 'COMPONENT', 'COMPONENTS', '模块', '组件'],
#             'max_chart_items': 15,
#             'max_pie_items': 8
#         },
#         'files': {
#             'csv_prefix': 'jira_issues',
#             'excel_prefix': 'jira_analysis'
#         }
#     }
    
#     try:
#         with open(config_file, 'r', encoding='utf-8') as file:
#             user_config = yaml.safe_load(file)
        
#         # 合并用户配置和默认配置
#         config = merge_config(default_config, user_config)
#         print(f"成功加载配置文件: {config_file}")
#         return config
#     except FileNotFoundError:
#         print(f"配置文件 {config_file} 不存在，使用默认配置")
#         # 保存默认配置文件
#         with open(config_file, 'w', encoding='utf-8') as file:
#             yaml.dump(default_config, file, default_flow_style=False, allow_unicode=True)
#         print(f"已创建默认配置文件: {config_file}")
#         return default_config
#     except Exception as e:
#         print(f"加载配置文件时出错: {e}，使用默认配置")
#         return default_config

# def merge_config(default_config, user_config):
#     """递归合并默认配置和用户配置"""
#     if not isinstance(user_config, dict):
#         return default_config
    
#     result = default_config.copy()
    
#     for key, value in user_config.items():
#         if key in result and isinstance(result[key], dict) and isinstance(value, dict):
#             result[key] = merge_config(result[key], value)
#         else:
#             result[key] = value
    
#     return result

# def get_credentials():
#     """使用GUI弹窗获取用户名和密码"""
#     try:
#         root = tk.Tk()
#         root.withdraw()  # 隐藏主窗口
        
#         # 创建自定义对话框
#         class CredentialsDialog(simpledialog.Dialog):
#             def __init__(self, parent, title=None):
#                 self.username = None
#                 self.password = None
#                 super().__init__(parent, title)
            
#             def body(self, master):
#                 tk.Label(master, text="Jira用户名:").grid(row=0, sticky="w")
#                 tk.Label(master, text="Jira密码:").grid(row=1, sticky="w")
                
#                 self.e1 = tk.Entry(master, width=30)
#                 self.e2 = tk.Entry(master, width=30, show="*")
                
#                 self.e1.grid(row=0, column=1, padx=5, pady=5)
#                 self.e2.grid(row=1, column=1, padx=5, pady=5)
                
#                 return self.e1  # 初始焦点
            
#             def apply(self):
#                 self.username = self.e1.get()
#                 self.password = self.e2.get()
        
#         # 显示对话框
#         dialog = CredentialsDialog(root, "Jira登录凭据")
        
#         # 销毁根窗口
#         root.destroy()
        
#         return dialog.username, dialog.password
        
#     except Exception as e:
#         print(f"GUI不可用，使用控制台输入: {e}")
#         username = input("请输入Jira用户名: ")
#         password = getpass.getpass("请输入Jira密码: ")
#         return username, password

# def download_with_progress(session, url, auth, timeout, filename):
#     """带进度显示的下载函数"""
#     response = session.get(url, auth=auth, timeout=timeout, stream=True)
#     response.raise_for_status()
    
#     total_size = int(response.headers.get('content-length', 0))
#     block_size = 8192
#     downloaded_size = 0
    
#     with open(filename, 'wb') as f:
#         for data in response.iter_content(block_size):
#             downloaded_size += len(data)
#             f.write(data)
            
#             # 显示下载进度
#             if total_size > 0:
#                 percent = (downloaded_size / total_size) * 100
#                 sys.stdout.write(f"\r下载进度: {downloaded_size}/{total_size} bytes ({percent:.1f}%)")
#                 sys.stdout.flush()
    
#     if total_size > 0:
#         print()  # 换行
    
#     return downloaded_size

# def download_jira_csv(config):
#     """下载Jira CSV文件"""
#     # 从配置获取URL，使用安全访问方式
#     url = config.get('jira', {}).get('url')
#     if not url:
#         print("错误：配置文件中缺少Jira URL")
#         return None
    
#     # 获取用户名和密码
#     username, password = get_credentials()
    
#     if not username or not password:
#         print("错误：未提供用户名或密码")
#         return None
    
#     # 创建会话并设置认证
#     session = requests.Session()
    
#     # 生成带时间戳的文件名，使用安全访问方式
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M")
#     csv_prefix = config.get('files', {}).get('csv_prefix', 'jira_issues')
#     csv_filename = f'{csv_prefix}_{timestamp}.csv'
    
#     # 设置请求头
#     headers = {
#         'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
#         'Accept': 'text/csv, */*',
#         'Accept-Encoding': 'gzip, deflate, br',
#         'Connection': 'keep-alive'
#     }
    
#     # 认证信息
#     auth = HTTPBasicAuth(username, password)
    
#     # 重试机制，使用安全访问方式
#     max_retries = config.get('jira', {}).get('retries', 3)
#     retry_delay = config.get('jira', {}).get('retry_delay', 5)
#     timeout = config.get('jira', {}).get('timeout', 60)
    
#     for attempt in range(max_retries):
#         try:
#             print(f"尝试连接Jira服务器 (尝试 {attempt + 1}/{max_retries})...")
#             print(f"URL: {url}")
            
#             # 首先尝试HEAD请求检查连接和认证
#             head_response = session.head(
#                 url,
#                 auth=auth,
#                 headers=headers,
#                 timeout=10
#             )
            
#             print(f"HEAD请求状态码: {head_response.status_code}")
            
#             if head_response.status_code == 401:
#                 print("认证失败：用户名或密码不正确")
#                 print("提示：如果启用了双重验证，请使用API令牌而不是密码")
#                 return None
#             elif head_response.status_code == 403:
#                 print("访问被拒绝：您没有权限访问此资源")
#                 return None
#             elif head_response.status_code == 404:
#                 print("文件未找到：URL可能已失效或JQL查询有误")
#                 return None
            
#             # 执行带进度显示的下载
#             print("开始下载CSV文件...")
#             downloaded_size = download_with_progress(
#                 session, url, auth, timeout, csv_filename
#             )
            
#             print(f"CSV文件下载成功！保存为: {csv_filename}")
#             print(f"文件大小: {downloaded_size} 字节")
            
#             # 显示前几行内容预览
#             try:
#                 with open(csv_filename, 'r', encoding='utf-8-sig') as f:
#                     content_preview = [f.readline() for _ in range(5)]
#                 print("\n文件前5行预览:")
#                 for i, line in enumerate(content_preview):
#                     print(f"{i+1}: {line.strip()}")
#             except UnicodeDecodeError:
#                 print("注意：文件内容无法用UTF-8解码")
            
#             return csv_filename
            
#         except requests.exceptions.Timeout:
#             print(f"连接超时 (尝试 {attempt + 1}/{max_retries})")
#             if attempt < max_retries - 1:
#                 print(f"等待 {retry_delay} 秒后重试...")
#                 time.sleep(retry_delay)
#             else:
#                 print("连接超时：请检查网络连接或尝试使用VPN")
#         except requests.exceptions.ConnectionError:
#             print(f"连接错误 (尝试 {attempt + 1}/{max_retries})")
#             if attempt < max_retries - 1:
#                 print(f"等待 {retry_delay} 秒后重试...")
#                 time.sleep(retry_delay)
#             else:
#                 print("连接错误：无法连接到服务器，请检查网络连接")
#                 print("提示：公司内部资源可能需要VPN连接")
#         except requests.exceptions.HTTPError as e:
#             print(f"HTTP错误: {e}")
#             if e.response.status_code == 401:
#                 print("认证失败：用户名或密码不正确")
#             elif e.response.status_code == 403:
#                 print("访问被拒绝：您没有权限访问此资源")
#             elif e.response.status_code == 404:
#                 print("文件未找到：URL可能已失效")
#             else:
#                 print(f"HTTP错误状态码: {e.response.status_code}")
#             return None
#         except requests.exceptions.RequestException as e:
#             print(f"网络请求错误: {e}")
#             if attempt < max_retries - 1:
#                 print(f"等待 {retry_delay} 秒后重试...")
#                 time.sleep(retry_delay)
#             else:
#                 print("网络请求失败，请检查网络连接和URL")
#         except Exception as e:
#             print(f"下载过程中发生未知错误: {e}")
#             return None
    
#     return None

# def analyze_jira_components(csv_file_path, config):
#     """
#     分析Jira CSV文件中的Component数量并生成带图表的Excel报告
#     """
#     # 生成带时间戳的输出文件名，使用安全访问方式
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M")
#     excel_prefix = config.get('files', {}).get('excel_prefix', 'jira_analysis')
#     output_excel_path = f'{excel_prefix}_{timestamp}.xlsx'
    
#     try:
#         # 检查CSV文件是否存在
#         if not os.path.exists(csv_file_path):
#             print(f"错误：找不到CSV文件 '{csv_file_path}'")
#             return False
        
#         # 读取CSV文件
#         print(f"正在读取文件: {csv_file_path}")
#         df = pd.read_csv(csv_file_path, encoding='utf-8-sig')
#         print(f"成功读取数据，共 {len(df)} 行，{len(df.columns)} 列")
        
#         # 显示所有列名，帮助识别Component列
#         print("\n数据列名:")
#         for i, col in enumerate(df.columns):
#             print(f"{i+1}: {col}")
        
#         # 从配置获取可能的Component列名，使用安全访问方式
#         component_col = None
#         possible_names = config.get('analysis', {}).get('component_column_names', 
#                                                        ['Component', 'Components', 'component', 'components', 
#                                                         'COMPONENT', 'COMPONENTS', '模块', '组件'])
        
#         for col in df.columns:
#             if col in possible_names:
#                 component_col = col
#                 break
                
#         # 如果没有自动识别到，尝试查找包含'component'的列名
#         if component_col is None:
#             for col in df.columns:
#                 if 'component' in col.lower():
#                     component_col = col
#                     break
        
#         # 如果仍然没有找到，让用户选择
#         if component_col is None:
#             print("\n未自动识别到Component列，请从以下列中选择:")
#             for i, col in enumerate(df.columns):
#                 print(f"{i+1}: {col}")
            
#             try:
#                 choice = int(input("请输入列号: ")) - 1
#                 if 0 <= choice < len(df.columns):
#                     component_col = df.columns[choice]
#                 else:
#                     print("无效选择，将使用第一列")
#                     component_col = df.columns[0]
#             except (ValueError, IndexError):
#                 print("输入无效，将使用第一列")
#                 component_col = df.columns[0]
            
#         print(f"\n使用列 '{component_col}' 进行Component统计")
        
#         # 统计Component数量
#         component_stats = analyze_components(df, component_col)
        
#         # 检查component_stats是否为空DataFrame
#         if component_stats is None or (hasattr(component_stats, 'empty') and component_stats.empty):
#             print("未找到有效的Component数据")
#             return False
        
#         # 创建Excel文件
#         create_excel_report(df, component_stats, component_col, output_excel_path, config)
        
#         print(f"\n分析完成！结果已保存到: {output_excel_path}")
#         return output_excel_path
        
#     except Exception as e:
#         print(f"分析过程中出错: {e}")
#         import traceback
#         traceback.print_exc()
#         return False

# def analyze_components(df, component_col):
#     """
#     分析Component数据并返回统计结果
#     """
#     all_components = []
    
#     # 检查列是否存在
#     if component_col not in df.columns:
#         print(f"错误：列 '{component_col}' 不存在于数据中")
#         return None
    
#     # 处理Component列（可能是字符串、列表或NaN）
#     for components in df[component_col].dropna():
#         if isinstance(components, str):
#             # 尝试解析字符串（可能是列表形式的字符串）
#             try:
#                 # 如果是类似 "['Comp1', 'Comp2']" 的格式
#                 if components.startswith('[') and components.endswith(']'):
#                     comp_list = ast.literal_eval(components)
#                     if isinstance(comp_list, list):
#                         all_components.extend(comp_list)
#                     else:
#                         all_components.append(str(comp_list))
#                 # 如果是分号分隔的
#                 elif ';' in components:
#                     all_components.extend([comp.strip() for comp in components.split(';') if comp.strip()])
#                 # 如果是逗号分隔的
#                 elif ',' in components:
#                     all_components.extend([comp.strip() for comp in components.split(',') if comp.strip()])
#                 else:
#                     all_components.append(components.strip())
#             except:
#                 # 如果解析失败，直接作为单个组件处理
#                 all_components.append(components.strip())
#         elif isinstance(components, list):
#             all_components.extend(components)
#         else:
#             all_components.append(str(components))
    
#     # 过滤空字符串
#     all_components = [comp for comp in all_components if comp and comp.strip()]
    
#     if not all_components:
#         print("没有找到有效的Component数据")
#         return None
    
#     # 统计数量
#     component_counts = Counter(all_components)
    
#     # 转换为DataFrame并排序
#     stats_df = pd.DataFrame({
#         'Component': list(component_counts.keys()),
#         'Count': list(component_counts.values())
#     }).sort_values('Count', ascending=False)
    
#     print(f"\n找到 {len(stats_df)} 个不同的Component")
#     print("\nComponent统计前10名:")
#     print(stats_df.head(10).to_string(index=False))
    
#     return stats_df

# def create_excel_report(df, component_stats, component_col, output_path, config):
#     """
#     创建包含统计数据和图表的Excel报告
#     """
#     # 创建Excel写入器
#     with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
#         # 1. 写入原始数据
#         df.to_excel(writer, sheet_name='原始数据', index=False)
        
#         # 2. 写入Component统计
#         component_stats.to_excel(writer, sheet_name='Component统计', index=False)
        
#         # 3. 创建汇总统计表
#         summary_data = {
#             '统计项': ['总Issue数', '有Component的Issue数', '无Component的Issue数', 
#                      '唯一Component数', '最多Component的Issue', '分析时间'],
#             '数值': [
#                 len(df),
#                 df[component_col].notna().sum(),
#                 df[component_col].isna().sum(),
#                 len(component_stats),
#                 f"{component_stats.iloc[0]['Component']} ({component_stats.iloc[0]['Count']}次)",
#                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#             ]
#         }
#         summary_df = pd.DataFrame(summary_data)
#         summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
        
#         # 获取工作簿和工作表以添加图表
#         workbook = writer.book
        
#         # 创建图表工作表
#         chart_sheet = workbook.create_sheet(title='图表')
        
#         # 生成图表
#         create_charts(component_stats, chart_sheet, workbook, config)
        
#         # 调整列宽
#         for sheet_name in writer.sheets:
#             worksheet = writer.sheets[sheet_name]
#             for column in worksheet.columns:
#                 max_length = 0
#                 column_letter = column[0].column_letter
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = min(max_length + 2, 50)
#                 worksheet.column_dimensions[column_letter].width = adjusted_width

# def create_charts(component_stats, chart_sheet, workbook, config):
#     """
#     在Excel中创建图表
#     """
#     from openpyxl.chart import BarChart, PieChart, Reference
    
#     # 从配置获取图表显示数量，使用安全访问方式
#     max_display = min(config.get('analysis', {}).get('max_chart_items', 15), len(component_stats))
#     pie_max = min(config.get('analysis', {}).get('max_pie_items', 8), len(component_stats))
    
#     display_data = component_stats.head(max_display)
#     pie_data = component_stats.head(pie_max)
    
#     # 将数据写入工作表用于图表
#     # 写入表头
#     chart_sheet['A1'] = 'Component'
#     chart_sheet['B1'] = 'Count'
    
#     # 写入数据
#     for i, (_, row) in enumerate(display_data.iterrows(), 2):
#         chart_sheet[f'A{i}'] = row['Component']
#         chart_sheet[f'B{i}'] = row['Count']
    
#     # 1. 创建条形图
#     bar_chart = BarChart()
#     bar_chart.title = f"Top {max_display} Components 统计"
#     bar_chart.style = 10
#     bar_chart.y_axis.title = '数量'
#     bar_chart.x_axis.title = 'Component'
    
#     # 数据引用
#     data = Reference(chart_sheet, min_col=2, min_row=1, max_row=max_display+1)
#     categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=max_display+1)
    
#     bar_chart.add_data(data, titles_from_data=True)
#     bar_chart.set_categories(categories)
    
#     # 将条形图添加到工作表
#     chart_sheet.add_chart(bar_chart, "D1")
    
#     # 2. 创建饼图
#     # 将数据写入工作表用于饼图
#     chart_sheet['A20'] = 'Component'
#     chart_sheet['B20'] = 'Count'
    
#     for i, (_, row) in enumerate(pie_data.iterrows(), 1):
#         chart_sheet.cell(row=i+20, column=1, value=row['Component'])
#         chart_sheet.cell(row=i+20, column=2, value=row['Count'])
    
#     pie_chart = PieChart()
#     pie_chart.title = f"Top {pie_max} Components 分布"
    
#     pie_data_ref = Reference(chart_sheet, min_col=2, min_row=21, max_row=20+pie_max)
#     pie_labels_ref = Reference(chart_sheet, min_col=1, min_row=21, max_row=20+pie_max)
    
#     pie_chart.add_data(pie_data_ref, titles_from_data=True)
#     pie_chart.set_categories(pie_labels_ref)
    
#     chart_sheet.add_chart(pie_chart, "D20")

# def main():
#     """
#     主函数 - 执行完整的流程
#     """
#     print("Jira CSV下载与分析工具")
#     print("=" * 50)
    
#     # 加载配置
#     config = load_config()
#     if not config:
#         print("无法加载配置，程序退出")
#         return
    
#     # 1. 下载CSV文件
#     print("步骤 1/2: 下载Jira CSV文件")
#     csv_file = download_jira_csv(config)
    
#     if not csv_file:
#         print("CSV文件下载失败，程序退出")
#         return
    
#     print(f"\n{'='*50}")
    
#     # 2. 分析CSV文件并生成Excel报告
#     print("步骤 2/2: 分析CSV文件并生成Excel报告")
#     excel_file = analyze_jira_components(csv_file, config)
    
#     if excel_file:
#         print(f"\n✅ 所有任务完成！")
#         print(f"📄 下载的CSV文件: {csv_file}")
#         print(f"📊 生成的Excel文件: {excel_file}")
#         print(f"\nExcel文件包含:")
#         print(f"   - 原始数据表")
#         print(f"   - Component统计表") 
#         print(f"   - 汇总统计表")
#         print(f"   - 条形图和饼图")
#     else:
#         print(f"\n❌ Excel报告生成失败")

# if __name__ == "__main__":
#     # 检查必要的库
#     try:
#         import openpyxl
#         import yaml
#     except ImportError:
#         print("缺少必要的库，正在安装...")
#         import subprocess
#         import sys
#         subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "matplotlib", "pandas", "requests", "PyYAML"])
#         print("库安装完成，请重新运行脚本")
#         exit()
    
#     main()
# # ```

# # ### 主要修复：

# # 1. **安全的配置访问**：
# #    - 使用 `config.get('section', {}).get('key', default_value)` 来安全地访问配置项
# #    - 为所有配置项提供了默认值，防止KeyError

# # 2. **改进的配置加载**：
# #    - 添加了 `merge_config` 函数来递归合并默认配置和用户配置
# #    - 即使配置文件缺少某些部分，也能正常工作

# # 3. **默认配置**：
# #    - 在代码中明确定义了完整的默认配置
# #    - 如果配置文件不存在或损坏，会自动使用默认配置

# # 4. **错误处理**：
# #    - 添加了更详细的错误信息
# #    - 即使配置不完整，程序也能继续运行

# # ### 使用说明：

# # 1. **首次运行**：
# #    - 如果 `config.yaml` 文件不存在，会自动创建包含完整默认配置的文件
# #    - 如果文件存在但缺少某些部分，程序会使用默认值填充缺失的部分

# # 2. **配置文件结构**：
# #    确保您的 `config.yaml` 文件包含以下结构：
# #    ```yaml
# #    jira:
# #      url: "您的Jira URL"
# #      timeout: 60
# #      retries: 3
# #      retry_delay: 5
   
# #    analysis:
# #      component_column_names: 
# #        - "Component"
# #        - "Components"
# #        - "component"
# #        - "components"
# #        - "COMPONENT"
# #        - "COMPONENTS"
# #        - "模块"
# #        - "组件"
# #      max_chart_items: 15
# #      max_pie_items: 8
   
# #    files:
# #      csv_prefix: "jira_issues"
# #      excel_prefix: "jira_analysis"
# #    ```

# # 3. **运行脚本**：
# #    ```bash
# #    python jira_download_analyze.py
# #    ```

# # 这个修复版本应该能够解决配置文件中缺少 `files` 部分的问题。如果您的配置文件已经存在但缺少某些部分，建议删除现有的配置文件，让程序重新生成一个完整的默认配置文件。

# 我将修改代码，完全去除预留的默认URL，确保URL只能从YAML配置文件中获取，如果找不到则报错：

# ```python
import requests
from requests.auth import HTTPBasicAuth
import tkinter as tk
from tkinter import simpledialog, messagebox
import pandas as pd
import matplotlib.pyplot as plt
from collections import Counter
import ast
import warnings
import os
from datetime import datetime
import getpass
import yaml
import time
import sys
import re

warnings.filterwarnings('ignore')

def load_config(config_file='config.yaml'):
    """从YAML文件加载配置，URL必须从配置文件中获取"""
    try:
        with open(config_file, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)
        
        if not config:
            print(f"❌ 错误：配置文件 {config_file} 为空")
            return None
        
        # 验证URL是否存在且格式正确
        url = config.get('jira', {}).get('url')
        if not url:
            print("❌ 错误：配置文件中缺少Jira URL")
            print("请在config.yaml文件的jira部分添加url字段")
            return None
        
        if not is_valid_url(url):
            print(f"❌ 错误：URL格式不正确: {url}")
            print("URL应该以http://或https://开头")
            return None
            
        print(f"✅ 配置加载成功")
        print(f"✅ URL验证通过: {url}")
        return config
        
    except FileNotFoundError:
        print(f"❌ 错误：找不到配置文件 {config_file}")
        print("请创建config.yaml文件，并包含以下结构:")
        print("""
jira:
  url: "您的Jira CSV下载URL"
  timeout: 60
  retries: 3
  retry_delay: 5

analysis:
  component_column_names: 
    - "Component"
    - "Components"
    - "component"
    - "components"
    - "COMPONENT"
    - "COMPONENTS"
    - "模块"
    - "组件"
  max_chart_items: 50
  max_pie_items: 20

files:
  csv_prefix: "jira_issues"
  excel_prefix: "jira_analysis"
""")
        return None
    except Exception as e:
        print(f"❌ 加载配置文件时出错: {e}")
        return None

def is_valid_url(url):
    """验证URL格式是否正确"""
    if not url or not isinstance(url, str):
        return False
    
    # 简单的URL格式验证
    url_pattern = re.compile(
        r'^(https?://)'  # http:// or https://
        r'([a-zA-Z0-9.-]+)'  # domain
        r'(\.[a-zA-Z]{2,})'  # top-level domain
        r'(/[^\s]*)?$'  # path
    )
    
    return bool(url_pattern.match(url))

def get_credentials():
    """使用GUI弹窗获取用户名和密码"""
    try:
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        
        # 创建自定义对话框
        class CredentialsDialog(simpledialog.Dialog):
            def __init__(self, parent, title=None):
                self.username = None
                self.password = None
                super().__init__(parent, title)
            
            def body(self, master):
                tk.Label(master, text="Jira用户名:").grid(row=0, sticky="w")
                tk.Label(master, text="Jira密码:").grid(row=1, sticky="w")
                
                self.e1 = tk.Entry(master, width=30)
                self.e2 = tk.Entry(master, width=30, show="*")
                
                self.e1.grid(row=0, column=1, padx=5, pady=5)
                self.e2.grid(row=1, column=1, padx=5, pady=5)
                
                return self.e1  # 初始焦点
            
            def apply(self):
                self.username = self.e1.get()
                self.password = self.e2.get()
        
        # 显示对话框
        dialog = CredentialsDialog(root, "Jira登录凭据")
        
        # 销毁根窗口
        root.destroy()
        
        return dialog.username, dialog.password
        
    except Exception as e:
        print(f"GUI不可用，使用控制台输入: {e}")
        username = input("请输入Jira用户名: ")
        password = getpass.getpass("请输入Jira密码: ")
        return username, password

def download_with_progress(session, url, auth, timeout, filename):
    """带进度显示的下载函数"""
    response = session.get(url, auth=auth, timeout=timeout, stream=True)
    response.raise_for_status()
    
    total_size = int(response.headers.get('content-length', 0))
    block_size = 8192
    downloaded_size = 0
    
    with open(filename, 'wb') as f:
        for data in response.iter_content(block_size):
            downloaded_size += len(data)
            f.write(data)
            
            # 显示下载进度
            if total_size > 0:
                percent = (downloaded_size / total_size) * 100
                sys.stdout.write(f"\r下载进度: {downloaded_size}/{total_size} bytes ({percent:.1f}%)")
                sys.stdout.flush()
    
    if total_size > 0:
        print()  # 换行
    
    return downloaded_size

def download_jira_csv(config):
    """下载Jira CSV文件"""
    # 从配置获取URL，使用安全访问方式
    url = config.get('jira', {}).get('url')
    if not url:
        print("❌ 错误：配置文件中缺少Jira URL")
        print("请在config.yaml文件的jira部分添加url字段")
        return None
    
    # 验证URL格式
    if not is_valid_url(url):
        print(f"❌ 错误：URL格式不正确: {url}")
        print("URL应该以http://或https://开头")
        return None
    
    # 获取用户名和密码
    username, password = get_credentials()
    
    if not username or not password:
        print("❌ 错误：未提供用户名或密码")
        return None
    
    # 创建会话并设置认证
    session = requests.Session()
    
    # 生成带时间戳的文件名，使用安全访问方式
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    csv_prefix = config.get('files', {}).get('csv_prefix', 'jira_issues')
    csv_filename = f'{csv_prefix}_{timestamp}.csv'
    
    # 设置请求头
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/csv, */*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive'
    }
    
    # 认证信息
    auth = HTTPBasicAuth(username, password)
    
    # 重试机制，使用安全访问方式
    max_retries = config.get('jira', {}).get('retries', 3)
    retry_delay = config.get('jira', {}).get('retry_delay', 5)
    timeout = config.get('jira', {}).get('timeout', 60)
    
    for attempt in range(max_retries):
        try:
            print(f"尝试连接Jira服务器 (尝试 {attempt + 1}/{max_retries})...")
            print(f"URL: {url}")
            
            # 首先尝试HEAD请求检查连接和认证
            head_response = session.head(
                url,
                auth=auth,
                headers=headers,
                timeout=10
            )
            
            print(f"HEAD请求状态码: {head_response.status_code}")
            
            if head_response.status_code == 401:
                print("❌ 认证失败：用户名或密码不正确")
                print("提示：如果启用了双重验证，请使用API令牌而不是密码")
                return None
            elif head_response.status_code == 403:
                print("❌ 访问被拒绝：您没有权限访问此资源")
                return None
            elif head_response.status_code == 404:
                print("❌ 文件未找到：URL可能已失效或JQL查询有误")
                print(f"请检查URL: {url}")
                return None
            elif head_response.status_code >= 500:
                print(f"❌ 服务器错误：状态码 {head_response.status_code}")
                print("请稍后重试或联系系统管理员")
                return None
            
            # 执行带进度显示的下载
            print("开始下载CSV文件...")
            downloaded_size = download_with_progress(
                session, url, auth, timeout, csv_filename
            )
            
            print(f"✅ CSV文件下载成功！保存为: {csv_filename}")
            print(f"文件大小: {downloaded_size} 字节")
            
            # 显示前几行内容预览
            try:
                with open(csv_filename, 'r', encoding='utf-8-sig') as f:
                    content_preview = [f.readline() for _ in range(5)]
                print("\n文件前5行预览:")
                for i, line in enumerate(content_preview):
                    print(f"{i+1}: {line.strip()}")
            except UnicodeDecodeError:
                print("注意：文件内容无法用UTF-8解码")
            
            return csv_filename
            
        except requests.exceptions.Timeout:
            print(f"❌ 连接超时 (尝试 {attempt + 1}/{max_retries})")
            if attempt < max_retries - 1:
                print(f"等待 {retry_delay} 秒后重试...")
                time.sleep(retry_delay)
            else:
                print("❌ 连接超时：请检查网络连接或尝试使用VPN")
        except requests.exceptions.ConnectionError:
            print(f"❌ 连接错误 (尝试 {attempt + 1}/{max_retries})")
            if attempt < max_retries - 1:
                print(f"等待 {retry_delay} 秒后重试...")
                time.sleep(retry_delay)
            else:
                print("❌ 连接错误：无法连接到服务器，请检查网络连接")
                print("提示：公司内部资源可能需要VPN连接")
        except requests.exceptions.HTTPError as e:
            print(f"❌ HTTP错误: {e}")
            if e.response.status_code == 401:
                print("认证失败：用户名或密码不正确")
            elif e.response.status_code == 403:
                print("访问被拒绝：您没有权限访问此资源")
            elif e.response.status_code == 404:
                print("文件未找到：URL可能已失效")
            else:
                print(f"HTTP错误状态码: {e.response.status_code}")
            return None
        except requests.exceptions.RequestException as e:
            print(f"❌ 网络请求错误: {e}")
            if attempt < max_retries - 1:
                print(f"等待 {retry_delay} 秒后重试...")
                time.sleep(retry_delay)
            else:
                print("❌ 网络请求失败，请检查网络连接和URL")
        except Exception as e:
            print(f"❌ 下载过程中发生未知错误: {e}")
            return None
    
    return None

def analyze_jira_components(csv_file_path, config):
    """
    分析Jira CSV文件中的Component数量并生成带图表的Excel报告
    """
    # 生成带时间戳的输出文件名，使用安全访问方式
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    excel_prefix = config.get('files', {}).get('excel_prefix', 'jira_analysis')
    output_excel_path = f'{excel_prefix}_{timestamp}.xlsx'
    
    try:
        # 检查CSV文件是否存在
        if not os.path.exists(csv_file_path):
            print(f"❌ 错误：找不到CSV文件 '{csv_file_path}'")
            return False
        
        # 读取CSV文件
        print(f"正在读取文件: {csv_file_path}")
        df = pd.read_csv(csv_file_path, encoding='utf-8-sig')
        print(f"✅ 成功读取数据，共 {len(df)} 行，{len(df.columns)} 列")
        
        # 显示所有列名，帮助识别Component列
        print("\n数据列名:")
        for i, col in enumerate(df.columns):
            print(f"{i+1}: {col}")
        
        # 从配置获取可能的Component列名，使用安全访问方式
        component_col = None
        possible_names = config.get('analysis', {}).get('component_column_names', 
                                                       ['Component', 'Components', 'component', 'components', 
                                                        'COMPONENT', 'COMPONENTS', '模块', '组件'])
        
        for col in df.columns:
            if col in possible_names:
                component_col = col
                break
                
        # 如果没有自动识别到，尝试查找包含'component'的列名
        if component_col is None:
            for col in df.columns:
                if 'component' in col.lower():
                    component_col = col
                    break
        
        # 如果仍然没有找到，让用户选择
        if component_col is None:
            print("\n未自动识别到Component列，请从以下列中选择:")
            for i, col in enumerate(df.columns):
                print(f"{i+1}: {col}")
            
            try:
                choice = int(input("请输入列号: ")) - 1
                if 0 <= choice < len(df.columns):
                    component_col = df.columns[choice]
                else:
                    print("无效选择，将使用第一列")
                    component_col = df.columns[0]
            except (ValueError, IndexError):
                print("输入无效，将使用第一列")
                component_col = df.columns[0]
            
        print(f"\n使用列 '{component_col}' 进行Component统计")
        
        # 统计Component数量
        component_stats = analyze_components(df, component_col)
        
        # 检查component_stats是否为空DataFrame
        if component_stats is None or (hasattr(component_stats, 'empty') and component_stats.empty):
            print("未找到有效的Component数据")
            return False
        
        # 创建Excel文件
        create_excel_report(df, component_stats, component_col, output_excel_path, config)
        
        print(f"\n✅ 分析完成！结果已保存到: {output_excel_path}")
        return output_excel_path
        
    except Exception as e:
        print(f"❌ 分析过程中出错: {e}")
        import traceback
        traceback.print_exc()
        return False

def analyze_components(df, component_col):
    """
    分析Component数据并返回统计结果
    """
    all_components = []
    
    # 检查列是否存在
    if component_col not in df.columns:
        print(f"❌ 错误：列 '{component_col}' 不存在于数据中")
        return None
    
    # 处理Component列（可能是字符串、列表或NaN）
    for components in df[component_col].dropna():
        if isinstance(components, str):
            # 尝试解析字符串（可能是列表形式的字符串）
            try:
                # 如果是类似 "['Comp1', 'Comp2']" 的格式
                if components.startswith('[') and components.endswith(']'):
                    comp_list = ast.literal_eval(components)
                    if isinstance(comp_list, list):
                        all_components.extend(comp_list)
                    else:
                        all_components.append(str(comp_list))
                # 如果是分号分隔的
                elif ';' in components:
                    all_components.extend([comp.strip() for comp in components.split(';') if comp.strip()])
                # 如果是逗号分隔的
                elif ',' in components:
                    all_components.extend([comp.strip() for comp in components.split(',') if comp.strip()])
                else:
                    all_components.append(components.strip())
            except:
                # 如果解析失败，直接作为单个组件处理
                all_components.append(components.strip())
        elif isinstance(components, list):
            all_components.extend(components)
        else:
            all_components.append(str(components))
    
    # 过滤空字符串
    all_components = [comp for comp in all_components if comp and comp.strip()]
    
    if not all_components:
        print("没有找到有效的Component数据")
        return None
    
    # 统计数量
    component_counts = Counter(all_components)
    
    # 转换为DataFrame并排序
    stats_df = pd.DataFrame({
        'Component': list(component_counts.keys()),
        'Count': list(component_counts.values())
    }).sort_values('Count', ascending=False)
    
    print(f"\n找到 {len(stats_df)} 个不同的Component")
    print("\nComponent统计前10名:")
    print(stats_df.head(10).to_string(index=False))
    
    return stats_df

def create_excel_report(df, component_stats, component_col, output_path, config):
    """
    创建包含统计数据和图表的Excel报告
    """
    # 创建Excel写入器
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # 1. 写入原始数据
        df.to_excel(writer, sheet_name='原始数据', index=False)
        
        # 2. 写入Component统计
        component_stats.to_excel(writer, sheet_name='Component统计', index=False)
        
        # 3. 创建汇总统计表
        summary_data = {
            '统计项': ['总Issue数', '有Component的Issue数', '无Component的Issue数', 
                     '唯一Component数', '最多Component的Issue', '分析时间'],
            '数值': [
                len(df),
                df[component_col].notna().sum(),
                df[component_col].isna().sum(),
                len(component_stats),
                f"{component_stats.iloc[0]['Component']} ({component_stats.iloc[0]['Count']}次)",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
        
        # 4. 创建百分比分布表
        create_percentage_table(component_stats, writer, config)
        
        # 获取工作簿和工作表以添加图表
        workbook = writer.book
        
        # 创建图表工作表
        chart_sheet = workbook.create_sheet(title='图表')
        
        # 生成图表（只生成条形图）
        create_charts(component_stats, chart_sheet, workbook, config)
        
        # 调整列宽
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

def create_percentage_table(component_stats, writer, config):
    """
    创建百分比分布表，替代饼图
    """
    # 计算前N个Component的百分比
    pie_max = min(config.get('analysis', {}).get('max_pie_items', 20), len(component_stats))
    top_components = component_stats.head(pie_max)
    
    # 计算百分比
    total_count = top_components['Count'].sum()
    top_components_with_percentage = top_components.copy()
    top_components_with_percentage['Percentage'] = (top_components_with_percentage['Count'] / total_count * 100).round(2)
    
    # 写入百分比分布表
    top_components_with_percentage.to_excel(writer, sheet_name='百分比分布', index=False)
    
    print(f"\n百分比分布表已创建，显示前 {pie_max} 个Component")
    print("百分比分布前10名:")
    print(top_components_with_percentage.head(10).to_string(index=False))

def create_charts(component_stats, chart_sheet, workbook, config):
    """
    在Excel中创建图表（只创建条形图，不创建饼图）
    """
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    
    # 使用全部数据，不限制显示数量
    max_display = len(component_stats)  # 使用全部数据
    
    display_data = component_stats.head(max_display)
    
    # 打印调试信息
    print(f"\n图表数据统计:")
    print(f"条形图显示 {len(display_data)} 个Component")
    
    # 将数据写入工作表用于图表
    # 写入表头
    chart_sheet['A1'] = 'Component'
    chart_sheet['B1'] = 'Count'
    
    # 写入柱状图数据
    for i, (_, row) in enumerate(display_data.iterrows(), 2):
        chart_sheet[f'A{i}'] = row['Component']
        chart_sheet[f'B{i}'] = row['Count']
    
    # 创建条形图
    bar_chart = BarChart()
    bar_chart.title = f"Components 统计 (共 {max_display} 个)"
    bar_chart.style = 10
    bar_chart.y_axis.title = '数量'
    bar_chart.x_axis.title = 'Component'
    
    # 设置条形图大小
    bar_chart.width = 30
    bar_chart.height = 15
    
    # 数据引用
    data = Reference(chart_sheet, min_col=2, min_row=1, max_row=max_display+1)
    categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=max_display+1)
    
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    
    # 添加数据标签（只显示数值，不显示任何文字）
    bar_chart.dataLabels = DataLabelList()
    bar_chart.dataLabels.showVal = True      # 显示数值
    bar_chart.dataLabels.showCatName = False # 不显示类别名称
    bar_chart.dataLabels.showSerName = False # 不显示系列名称
    bar_chart.dataLabels.showLegendKey = False # 不显示图例标示
    bar_chart.dataLabels.dLblPos = 'outEnd'  # 标签位置在柱子外部
    
    # 将条形图添加到工作表
    chart_sheet.add_chart(bar_chart, "D1")
    
    # 在图表工作表下方添加百分比数据说明
    note_row = max_display + 5
    chart_sheet[f'A{note_row}'] = "说明:"
    chart_sheet[f'A{note_row+1}'] = "1. 此图表显示所有Component的数量统计"
    chart_sheet[f'A{note_row+2}'] = "2. 每个柱子上方显示该Component的具体数量"
    chart_sheet[f'A{note_row+3}'] = "3. 详细百分比分布请查看'百分比分布'工作表"

def main():
    """
    主函数 - 执行完整的流程
    """
    print("Jira CSV下载与分析工具")
    print("=" * 50)
    
    # 加载配置
    config = load_config()
    if not config:
        print("❌ 配置加载失败，程序退出")
        return
    
    # 1. 下载CSV文件
    print("步骤 1/2: 下载Jira CSV文件")
    csv_file = download_jira_csv(config)
    
    if not csv_file:
        print("❌ CSV文件下载失败，程序退出")
        return
    
    print(f"\n{'='*50}")
    
    # 2. 分析CSV文件并生成Excel报告
    print("步骤 2/2: 分析CSV文件并生成Excel报告")
    excel_file = analyze_jira_components(csv_file, config)
    
    if excel_file:
        print(f"\n✅ 所有任务完成！")
        print(f"📄 下载的CSV文件: {csv_file}")
        print(f"📊 生成的Excel文件: {excel_file}")
        print(f"\nExcel文件包含:")
        print(f"   - 原始数据表")
        print(f"   - Component统计表") 
        print(f"   - 汇总统计表")
        print(f"   - 百分比分布表（替代饼图）")
        print(f"   - 条形图（显示全部Component数据）")
        print(f"   - 条形图只显示数字")
    else:
        print(f"\n❌ Excel报告生成失败")

if __name__ == "__main__":
    # 检查必要的库
    try:
        import openpyxl
        import yaml
    except ImportError:
        print("缺少必要的库，正在安装...")
        import subprocess
        import sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "matplotlib", "pandas", "requests", "PyYAML"])
        print("库安装完成，请重新运行脚本")
        exit()
    
    main()
# ```

# ### 主要修改：

# 1. **完全去除预留URL**：
#    - 移除了所有默认配置中的URL
#    - 不再自动创建默认配置文件

# 2. **严格的URL验证**：
#    - 配置文件必须存在且包含有效的URL
#    - 如果配置文件不存在，会显示详细的错误信息和配置文件结构示例
#    - 如果配置文件中缺少URL或URL格式不正确，会报错并退出

# 3. **清晰的错误提示**：
#    - 当配置文件不存在时，显示完整的配置文件结构示例
#    - 当URL缺失或格式不正确时，提供具体的错误原因和解决方案

# ### 配置文件要求：

# 现在您必须创建一个 `config.yaml` 文件，包含以下结构：

# ```yaml
# jira:
#   url: "您的Jira CSV下载URL"
#   timeout: 60
#   retries: 3
#   retry_delay: 5

# analysis:
#   component_column_names: 
#     - "Component"
#     - "Components"
#     - "component"
#     - "components"
#     - "COMPONENT"
#     - "COMPONENTS"
#     - "模块"
#     - "组件"
#   max_chart_items: 50
#   max_pie_items: 20

# files:
#   csv_prefix: "jira_issues"
#   excel_prefix: "jira_analysis"
# ```

# ### 使用说明：

# 1. **创建配置文件**：
#    - 在与脚本相同的目录下创建 `config.yaml` 文件
#    - 按照上述结构填写配置，特别是 `jira.url` 字段

# 2. **运行脚本**：
#    ```bash
#    python jira_download_analyze.py
#    ```

# 3. **错误处理**：
#    - 如果配置文件不存在，脚本会显示错误并退出
#    - 如果配置文件中缺少URL或URL格式不正确，脚本会显示错误并退出

# ### 注意事项：

# - 脚本现在完全依赖配置文件，没有内置的默认URL
# - 您必须提供有效的Jira CSV下载URL
# - 配置文件必须符合YAML格式要求

# 这样修改后，URL只能从YAML配置文件中获取，如果找不到或格式不正确，脚本会明确报错并退出，确保配置的正确性。