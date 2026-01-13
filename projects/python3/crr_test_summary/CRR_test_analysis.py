
# import re
# import pandas as pd
# import os
# import yaml
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows

# def load_config(config_file='config.yaml'):
#     """
#     加载YAML配置文件
#     """
#     try:
#         with open(config_file, 'r', encoding='utf-8') as file:
#             config = yaml.safe_load(file)
#         return config
#     except Exception as e:
#         print(f"Error loading config file {config_file}: {str(e)}")
#         # 返回默认配置
#         return {
#             'log_directory': '.',
#             'reuse_ip': []
#         }

# def analyze_uvm_log(file_path):
#     """
#     分析UVM日志文件，提取UVM_ERROR和UVM_FATAL数量
#     """
#     try:
#         with open(file_path, 'r', encoding='utf-8') as file:
#             content = file.read()
        
#         # 检查是否包含UVM_ERROR或UVM_FATAL
#         if 'UVM_ERROR' not in content and 'UVM_FATAL' not in content:
#             return None  # 忽略不包含这些关键字的文件
        
#         # 提取UVM_ERROR数量
#         error_match = re.search(r'UVM_ERROR\s*:\s*(\d+)', content)
#         error_count = int(error_match.group(1)) if error_match else 0
        
#         # 提取UVM_FATAL数量
#         fatal_match = re.search(r'UVM_FATAL\s*:\s*(\d+)', content)
#         fatal_count = int(fatal_match.group(1)) if fatal_match else 0
        
#         # 判断测试结果
#         if error_count == 0 and fatal_count == 0:
#             result = "PASS"
#         else:
#             result = "FAIL"
        
#         # 获取文件名
#         file_name = os.path.basename(file_path)
        
#         return {
#             'File Name': file_name,
#             'UVM_ERROR': error_count,
#             'UVM_FATAL': fatal_count,
#             'Result': result,
#             'Full Path': file_path  # 保存完整路径用于调试
#         }
    
#     except Exception as e:
#         print(f"Error processing file {file_path}: {str(e)}")
#         return {
#             'File Name': os.path.basename(file_path),
#             'UVM_ERROR': -1,
#             'UVM_FATAL': -1,
#             'Result': 'ERROR',
#             'Full Path': file_path
#         }

# def check_reuse_ip_match(file_name, reuse_ip_list):
#     """
#     检查文件名是否匹配reuse IP列表中的任何一项
#     """
#     for ip in reuse_ip_list:
#         if ip in file_name:
#             return ip
#     return None

# def generate_excel_report(config, output_file='crr_test_report.xlsx'):
#     """
#     生成Excel测试报告，包含颜色标记和多个汇总sheet
#     """
#     directory_path = config.get('log_directory', '.')
#     reuse_ip_list = config.get('reuse_ip', [])
    
#     # 获取目录下所有.log文件
#     log_files = []
#     for file in os.listdir(directory_path):
#         if file.endswith('.log'):
#             log_files.append(os.path.join(directory_path, file))
    
#     if not log_files:
#         print(f"在目录 {directory_path} 中未找到.log文件")
#         return None
    
#     print(f"找到 {len(log_files)} 个日志文件")
    
#     # 分析所有日志文件
#     results = []
#     skipped_files = 0
    
#     for log_file in log_files:
#         result = analyze_uvm_log(log_file)
#         if result is None:
#             skipped_files += 1
#             print(f"跳过文件 (不包含UVM_ERROR/FATAL): {os.path.basename(log_file)}")
#             continue
            
#         # 检查是否匹配reuse IP
#         matched_ip = check_reuse_ip_match(result['File Name'], reuse_ip_list)
#         result['Matched IP'] = matched_ip if matched_ip else ""
        
#         results.append(result)
#         print(f"分析完成: {result['File Name']} - {result['Result']}")
    
#     print(f"跳过了 {skipped_files} 个不包含UVM_ERROR/FATAL的文件")
    
#     if not results:
#         print("没有找到包含UVM_ERROR/FATAL的有效日志文件")
#         return None
    
#     # 创建DataFrame
#     df = pd.DataFrame(results)
    
#     # 创建Excel工作簿
#     wb = Workbook()
    
#     # 删除默认创建的sheet
#     wb.remove(wb.active)
    
#     # 创建所有结果的sheet
#     ws_all = wb.create_sheet("All Results")
    
#     # 将数据写入sheet
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
#         for c_idx, value in enumerate(row, 1):
#             ws_all.cell(row=r_idx, column=c_idx, value=value)
    
#     # 设置列宽
#     for column in ws_all.columns:
#         max_length = 0
#         column_letter = column[0].column_letter
#         for cell in column:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = (max_length + 2)
#         ws_all.column_dimensions[column_letter].width = adjusted_width
    
#     # 设置颜色填充
#     green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
#     red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
#     # 应用颜色到结果列
#     for row in range(2, len(df) + 2):  # 从第2行开始（跳过标题）
#         result_cell = ws_all.cell(row=row, column=4)  # 第4列是Result
#         if result_cell.value == "PASS":
#             for col in range(1, 6):  # 对前5列应用颜色
#                 ws_all.cell(row=row, column=col).fill = green_fill
#         elif result_cell.value == "FAIL":
#             for col in range(1, 6):  # 对前5列应用颜色
#                 ws_all.cell(row=row, column=col).fill = red_fill
    
#     # 创建失败汇总sheet
#     fail_df = df[df['Result'] == 'FAIL']
#     if not fail_df.empty:
#         ws_fail = wb.create_sheet("Fail Summary")
        
#         # 将失败数据写入sheet
#         for r_idx, row in enumerate(dataframe_to_rows(fail_df, index=False, header=True), 1):
#             for c_idx, value in enumerate(row, 1):
#                 ws_fail.cell(row=r_idx, column=c_idx, value=value)
        
#         # 设置列宽
#         for column in ws_fail.columns:
#             max_length = 0
#             column_letter = column[0].column_letter
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = (max_length + 2)
#             ws_fail.column_dimensions[column_letter].width = adjusted_width
        
#         # 对失败汇总应用红色
#         for row in range(2, len(fail_df) + 2):
#             for col in range(1, 6):
#                 ws_fail.cell(row=row, column=col).fill = red_fill
    
#     # 创建Reuse IP失败汇总sheet
#     if reuse_ip_list:
#         reuse_ip_fail_df = fail_df[fail_df['Matched IP'] != ""]
#         if not reuse_ip_fail_df.empty:
#             ws_reuse_ip_fail = wb.create_sheet("Reuse IP Fail Summary")
            
#             # 将Reuse IP失败数据写入sheet
#             for r_idx, row in enumerate(dataframe_to_rows(reuse_ip_fail_df, index=False, header=True), 1):
#                 for c_idx, value in enumerate(row, 1):
#                     ws_reuse_ip_fail.cell(row=r_idx, column=c_idx, value=value)
            
#             # 设置列宽
#             for column in ws_reuse_ip_fail.columns:
#                 max_length = 0
#                 column_letter = column[0].column_letter
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 ws_reuse_ip_fail.column_dimensions[column_letter].width = adjusted_width
            
#             # 对Reuse IP失败汇总应用红色
#             for row in range(2, len(reuse_ip_fail_df) + 2):
#                 for col in range(1, 6):
#                     ws_reuse_ip_fail.cell(row=row, column=col).fill = red_fill
    
#     # 保存Excel文件
#     wb.save(output_file)
#     print(f"Excel报告已生成: {output_file}")
    
#     # 打印汇总信息
#     total_files = len(results)
#     pass_count = sum(1 for r in results if r['Result'] == 'PASS')
#     fail_count = sum(1 for r in results if r['Result'] == 'FAIL')
#     error_count = sum(1 for r in results if r['Result'] == 'ERROR')
#     reuse_ip_count = sum(1 for r in results if r['Matched IP'] != "")
    
#     print(f"\n测试结果汇总:")
#     print(f"总文件数: {total_files}")
#     print(f"PASS: {pass_count}")
#     print(f"FAIL: {fail_count}")
#     if error_count > 0:
#         print(f"ERROR: {error_count}")
#     if reuse_ip_list:
#         print(f"匹配Reuse IP的文件: {reuse_ip_count}")
    
#     return df

# # 使用示例
# if __name__ == "__main__":
#     # 加载配置文件
#     config = load_config()
    
#     # 生成Excel报告
#     report_df = generate_excel_report(config)
    
#     if report_df is not None:
#         # 显示报告内容
#         print("\n生成的报告内容:")
#         print(report_df[['File Name', 'UVM_ERROR', 'UVM_FATAL', 'Result', 'Matched IP']])

import re
import pandas as pd
import os
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def load_config(config_file='config.yaml'):
    """
    Load YAML configuration file
    """
    try:
        with open(config_file, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)
        return config
    except Exception as e:
        print(f"Error loading config file {config_file}: {str(e)}")
        # Return default configuration
        return {
            'log_directory': '.',
            'reuse_ip': []
        }

def analyze_uvm_log(file_path):
    """
    Analyze UVM log file to extract UVM_ERROR and UVM_FATAL counts
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # Check if file contains UVM_ERROR or UVM_FATAL
        if 'UVM_ERROR' not in content and 'UVM_FATAL' not in content:
            return None  # Ignore files without these keywords
        
        # Extract UVM_ERROR count
        error_match = re.search(r'UVM_ERROR\s*:\s*(\d+)', content)
        error_count = int(error_match.group(1)) if error_match else 0
        
        # Extract UVM_FATAL count
        fatal_match = re.search(r'UVM_FATAL\s*:\s*(\d+)', content)
        fatal_count = int(fatal_match.group(1)) if fatal_match else 0
        
        # Determine test result
        if error_count == 0 and fatal_count == 0:
            result = "PASS"
        else:
            result = "FAIL"
        
        # Get filename
        file_name = os.path.basename(file_path)
        
        return {
            'File Name': file_name,
            'UVM_ERROR': error_count,
            'UVM_FATAL': fatal_count,
            'Result': result,
            'Full Path': file_path  # Save full path for debugging
        }
    
    except Exception as e:
        print(f"Error processing file {file_path}: {str(e)}")
        return {
            'File Name': os.path.basename(file_path),
            'UVM_ERROR': -1,
            'UVM_FATAL': -1,
            'Result': 'ERROR',
            'Full Path': file_path
        }

def check_reuse_ip_match(file_name, reuse_ip_list):
    """
    Check if filename matches any item in the reuse IP list
    """
    for ip in reuse_ip_list:
        if ip in file_name:
            return ip
    return None

def generate_excel_report(config):
    """
    Generate Excel test report with color coding and multiple summary sheets
    """
    directory_path = config.get('log_directory', '.')
    reuse_ip_list = config.get('reuse_ip', [])
    
    # Get all .log files in directory
    log_files = []
    for file in os.listdir(directory_path):
        if file.endswith('.log'):
            log_files.append(os.path.join(directory_path, file))
    
    if not log_files:
        print(f"No .log files found in directory {directory_path}")
        return None
    
    print(f"Found {len(log_files)} log files")
    
    # Analyze all log files
    results = []
    skipped_files = 0
    
    for log_file in log_files:
        result = analyze_uvm_log(log_file)
        if result is None:
            skipped_files += 1
            print(f"Skipped file (no UVM_ERROR/FATAL): {os.path.basename(log_file)}")
            continue
            
        # Check if matches reuse IP
        matched_ip = check_reuse_ip_match(result['File Name'], reuse_ip_list)
        result['Matched IP'] = matched_ip if matched_ip else ""
        
        results.append(result)
        print(f"Analysis completed: {result['File Name']} - {result['Result']}")
    
    print(f"Skipped {skipped_files} files without UVM_ERROR/FATAL")
    
    if not results:
        print("No valid log files with UVM_ERROR/FATAL found")
        return None
    
    # Create DataFrame
    df = pd.DataFrame(results)
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default created sheet
    wb.remove(wb.active)
    
    # Create sheet for all results
    ws_all = wb.create_sheet("All Results")
    
    # Write data to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_all.cell(row=r_idx, column=c_idx, value=value)
    
    # Set column widths
    for column in ws_all.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_all.column_dimensions[column_letter].width = adjusted_width
    
    # Set color fills
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Apply colors to result column
    for row in range(2, len(df) + 2):  # Start from row 2 (skip header)
        result_cell = ws_all.cell(row=row, column=4)  # Column 4 is Result
        if result_cell.value == "PASS":
            for col in range(1, 6):  # Apply to first 5 columns
                ws_all.cell(row=row, column=col).fill = green_fill
        elif result_cell.value == "FAIL":
            for col in range(1, 6):  # Apply to first 5 columns
                ws_all.cell(row=row, column=col).fill = red_fill
    
    # Create failure summary sheet
    fail_df = df[df['Result'] == 'FAIL']
    if not fail_df.empty:
        ws_fail = wb.create_sheet("Fail Summary")
        
        # Write failure data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(fail_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_fail.cell(row=r_idx, column=c_idx, value=value)
        
        # Set column widths
        for column in ws_fail.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_fail.column_dimensions[column_letter].width = adjusted_width
        
        # Apply red to failure summary
        for row in range(2, len(fail_df) + 2):
            for col in range(1, 6):
                ws_fail.cell(row=row, column=col).fill = red_fill
    
    # Create Reuse IP failure summary sheet
    if reuse_ip_list:
        reuse_ip_fail_df = fail_df[fail_df['Matched IP'] != ""]
        if not reuse_ip_fail_df.empty:
            ws_reuse_ip_fail = wb.create_sheet("Reuse IP Fail Summary")
            
            # Write Reuse IP failure data to sheet
            for r_idx, row in enumerate(dataframe_to_rows(reuse_ip_fail_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_reuse_ip_fail.cell(row=r_idx, column=c_idx, value=value)
            
            # Set column widths
            for column in ws_reuse_ip_fail.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_reuse_ip_fail.column_dimensions[column_letter].width = adjusted_width
            
            # Apply red to Reuse IP failure summary
            for row in range(2, len(reuse_ip_fail_df) + 2):
                for col in range(1, 6):
                    ws_reuse_ip_fail.cell(row=row, column=col).fill = red_fill
    
    # Generate timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_file = f"crr_test_report_{timestamp}.xlsx"
    
    # Save Excel file
    wb.save(output_file)
    print(f"Excel report generated: {output_file}")
    
    # Print summary information
    total_files = len(results)
    pass_count = sum(1 for r in results if r['Result'] == 'PASS')
    fail_count = sum(1 for r in results if r['Result'] == 'FAIL')
    error_count = sum(1 for r in results if r['Result'] == 'ERROR')
    reuse_ip_count = sum(1 for r in results if r['Matched IP'] != "")
    
    print(f"\nTest Results Summary:")
    print(f"Total files: {total_files}")
    print(f"PASS: {pass_count}")
    print(f"FAIL: {fail_count}")
    if error_count > 0:
        print(f"ERROR: {error_count}")
    if reuse_ip_list:
        print(f"Files matching Reuse IP: {reuse_ip_count}")
    
    return df, output_file

# Main execution
if __name__ == "__main__":
    # Load configuration file
    config = load_config()
    
    # Generate Excel report
    result = generate_excel_report(config)
    
    if result is not None:
        report_df, output_file = result
        # Display report content
        print("\nGenerated report content:")
        print(report_df[['File Name', 'UVM_ERROR', 'UVM_FATAL', 'Result', 'Matched IP']])

