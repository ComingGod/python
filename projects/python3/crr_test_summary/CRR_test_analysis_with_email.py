
import re
import os
import sys
import yaml
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import matplotlib.pyplot as plt

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders

# =========================================================
# 1) CSV -> YAML
# =========================================================
def generate_yaml_from_csv(csv_file, log_directory):
    if not os.path.isfile(csv_file):
        print(f"Error: CSV file '{csv_file}' not found.")
        sys.exit(1)
    df = pd.read_csv(csv_file)
    required_cols = {"IP_Name", "Owner", "IP type", "reuse"}
    if not required_cols.issubset(df.columns):
        print(f"Error: CSV file must contain columns: {required_cols}")
        sys.exit(1)
    config = {"log_directory": log_directory, "reuse_ip": [], "new_ip": []}
    for _, row in df.iterrows():
        ip_entry = {
            "name": str(row["IP_Name"]),
            "ip_type": str(row["IP type"]),
            "owner": str(row["Owner"])
        }
        if str(row["reuse"]).strip().lower() == "yes":
            config["reuse_ip"].append(ip_entry)
        else:
            config["new_ip"].append(ip_entry)
    yaml_file = os.path.join(os.getcwd(), os.path.splitext(os.path.basename(csv_file))[0] + "_config.yaml")
    with open(yaml_file, "w", encoding="utf-8") as f:
        yaml.dump(config, f, sort_keys=False, allow_unicode=True)
    print(f"YAML file generated: {yaml_file}")
    return yaml_file

# =========================================================
# 2) Load YAML
# =========================================================
def load_config(config_file):
    try:
        with open(config_file, 'r', encoding='utf-8') as file:
            return yaml.safe_load(file)
    except Exception as e:
        print(f"Error loading config file: {str(e)}")
        return {'log_directory': '.', 'reuse_ip': [], 'new_ip': []}

# =========================================================
# 3) Parse UVM logs
# =========================================================
def analyze_uvm_log(file_path):
    """Extract PASS/FAIL/NONE by counting UVM_ERROR/UVM_FATAL."""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        if 'UVM_ERROR' not in content and 'UVM_FATAL' not in content:
            return "NONE"
        error_match = re.search(r'UVM_ERROR\s*:\s*(\d+)', content)
        fatal_match = re.search(r'UVM_FATAL\s*:\s*(\d+)', content)
        error_count = int(error_match.group(1)) if error_match else 0
        fatal_count = int(fatal_match.group(1)) if fatal_match else 0
        return "PASS" if error_count == 0 and fatal_count == 0 else "FAIL"
    except Exception:
        return "NONE"

def get_result_priority(result):
    return 1 if result == "FAIL" else (2 if result == "NONE" else 3)

# =========================================================
# 4) Excel report + stats
# =========================================================
def generate_excel_report(config):
    print("Processing...")
    directory_path = config.get('log_directory', '.')
    reuse_ip_list = config.get('reuse_ip', [])
    new_ip_list = config.get('new_ip', [])
    results_dir = 'results'
    os.makedirs(results_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    ws_results = wb.create_sheet("IP Test Results")

    headers = ["IP Name", "IP Category", "IP Type", "Owner", "Reg Access", "Reset Check"]
    for col, header in enumerate(headers, 1):
        ws_results.cell(row=1, column=col, value=header)

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    stats = {
        'total': {'cases': 0, 'pass': 0},
        'reuse_ip': {'cases': 0, 'pass': 0},
        'new_ip': {'cases': 0, 'pass': 0},
        'ip_types': {}
    }

    # Build IP list
    all_results = []
    for ip_info in reuse_ip_list:
        all_results.append((ip_info['name'], ip_info.get('ip_type', 'Unknown'),
                            ip_info.get('owner', 'Unknown'), 'Reuse'))
    for ip_info in new_ip_list:
        all_results.append((ip_info['name'], ip_info.get('ip_type', 'Unknown'),
                            ip_info.get('owner', 'Unknown'), 'New'))

    # Process each IP
    ip_data = []
    for ip_name, ip_type, owner, ip_category in all_results:
        reg_access_result = "NONE"
        reset_check_result = "NONE"
        for test_type in ["reg_access", "reset_check"]:
            matching_files = [
                file for file in os.listdir(directory_path)
                if (
                    file.lower().startswith('mdv_rtl') and
                    file.lower().endswith('.log') and
                    ip_name.lower() in file.lower() and
                    test_type in file.lower()
                )
            ]
            if matching_files:
                file_path = os.path.join(directory_path, matching_files[0])
                result = analyze_uvm_log(file_path)
                if test_type == "reg_access":
                    reg_access_result = result
                else:
                    reset_check_result = result

        priority = min(get_result_priority(reg_access_result),
                       get_result_priority(reset_check_result))

        ip_data.append({
            'name': ip_name,
            'category': ip_category,
            'type': ip_type,
            'owner': owner,
            'reg_access': reg_access_result,
            'reset_check': reset_check_result,
            'priority': priority
        })

        # Stats
        stats['total']['cases'] += 2
        if reg_access_result == "PASS": stats['total']['pass'] += 1
        if reset_check_result == "PASS": stats['total']['pass'] += 1

        cat_key = 'reuse_ip' if ip_category == 'Reuse' else 'new_ip'
        stats[cat_key]['cases'] += 2
        if reg_access_result == "PASS": stats[cat_key]['pass'] += 1
        if reset_check_result == "PASS": stats[cat_key]['pass'] += 1

        if ip_type not in stats['ip_types']:
            stats['ip_types'][ip_type] = {'cases': 0, 'pass': 0}
        stats['ip_types'][ip_type]['cases'] += 2
        if reg_access_result == "PASS": stats['ip_types'][ip_type]['pass'] += 1
        if reset_check_result == "PASS": stats['ip_types'][ip_type]['pass'] += 1

    # Sort by priority
    ip_data.sort(key=lambda x: x['priority'])

    # Write rows
    row = 2
    for ip in ip_data:
        ws_results.cell(row=row, column=1, value=ip['name'])
        ws_results.cell(row=row, column=2, value=ip['category'])
        ws_results.cell(row=row, column=3, value=ip['type'])
        ws_results.cell(row=row, column=4, value=ip['owner'])
        ws_results.cell(row=row, column=5, value=ip['reg_access'])
        ws_results.cell(row=row, column=6, value=ip['reset_check'])

        ws_results.cell(row=row, column=5).fill = (
            green_fill if ip['reg_access'] == "PASS"
            else red_fill if ip['reg_access'] == "FAIL"
            else yellow_fill
        )
        ws_results.cell(row=row, column=6).fill = (
            green_fill if ip['reset_check'] == "PASS"
            else red_fill if ip['reset_check'] == "FAIL"
            else yellow_fill
        )
        row += 1

    # Auto column width
    for column in ws_results.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws_results.column_dimensions[column[0].column_letter].width = max_length + 2

    # Save report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    excel_file = os.path.join(results_dir, f"crr_result_analysis_{timestamp}.xlsx")
    wb.save(excel_file)
    print(f"Report saved to: {excel_file}")

    # Build HTML + charts
    html_file, email_html, inline_images = build_html_and_email_reports(stats, excel_file, results_dir)
    print(f"HTML report saved to: {html_file}")

    return excel_file, html_file, email_html, inline_images

# =========================================================
# 5) HTML + charts (local HTML + email HTML with CID)
# =========================================================
def build_html_and_email_reports(stats, excel_file, results_dir):
    os.makedirs(results_dir, exist_ok=True)

    overall_png = os.path.join(results_dir, 'overall_chart.png')
    category_png = os.path.join(results_dir, 'category_chart.png')
    type_png = os.path.join(results_dir, 'type_chart.png')

    # Charts
    fig1, ax1 = plt.subplots()
    ax1.pie([stats['total']['pass'], stats['total']['cases'] - stats['total']['pass']],
            labels=['PASS', 'FAIL'], autopct='%1.1f%%', colors=['#4CAF50', '#F44336'])
    ax1.set_title('Overall Test Results')
    fig1.savefig(overall_png, bbox_inches='tight')
    plt.close(fig1)

    fig2, ax2 = plt.subplots()
    ax2.pie([stats['reuse_ip']['pass'], stats['reuse_ip']['cases'] - stats['reuse_ip']['pass'],
             stats['new_ip']['pass'], stats['new_ip']['cases'] - stats['new_ip']['pass']],
            labels=['Reuse PASS', 'Reuse FAIL', 'New PASS', 'New FAIL'], autopct='%1.1f%%')
    ax2.set_title('PASS/FAIL by IP Category')
    fig2.savefig(category_png, bbox_inches='tight')
    plt.close(fig2)

    fig3, ax3 = plt.subplots()
    ip_types = list(stats['ip_types'].keys())
    pass_counts = [stats['ip_types'][t]['pass'] for t in ip_types]
    fail_counts = [stats['ip_types'][t]['cases'] - stats['ip_types'][t]['pass'] for t in ip_types]
    x = range(len(ip_types))
    ax3.bar(x, pass_counts, width=0.4, label='PASS', color='#4CAF50')
    ax3.bar(x, fail_counts, width=0.4, bottom=pass_counts, label='FAIL', color='#F44336')
    ax3.set_xticks(x)
    ax3.set_xticklabels(ip_types, rotation=45, ha='right')
    ax3.set_ylabel('Number of Cases')
    ax3.set_title('PASS/FAIL by IP Type')
    ax3.legend()
    fig3.tight_layout()
    fig3.savefig(type_png, bbox_inches='tight')
    plt.close(fig3)

    # Local HTML (browser preview: images in same dir by basename)
    css = """body { font-family: Arial, Helvetica, sans-serif; }
h2 { margin-bottom: 0.4rem; }
h3 { margin-top: 1.0rem; margin-bottom: 0.4rem; }
img { max-width: 920px; height: auto; }"""
    total_cases = stats['total']['cases']
    pass_rate = (stats['total']['pass'] / total_cases) * 100 if total_cases > 0 else 0.0

    local_html = (
        "<html><head><meta charset='utf-8'>"
        "<title>CRR Test Analysis Summary</title>"
        f"<style>{css}</style></head><body>"
        "<h2>CRR Test Analysis Summary</h2>"
        "<h3>Overall Statistics</h3>"
        f"<p>Total Cases: {total_cases}<br>"
        f"   PASS: {stats['total']['pass']}<br>"
        f"   FAIL: {stats['total']['cases'] - stats['total']['pass']}<br>"
        f"   Pass Rate: {pass_rate:.1f}%</p>"
        "<verall_chart.png<br>"
        "<h3>Category Statistics</h3>"
        f"<p>Reuse IP - Cases: {stats['reuse_ip']['cases']}, PASS: {stats['reuse_ip']['pass']}, "
        f"FAIL: {stats['reuse_ip']['cases'] - stats['reuse_ip']['pass']}<br>"
        f"New IP - Cases: {stats['new_ip']['cases']}, PASS: {stats['new_ip']['pass']}, "
        f"FAIL: {stats['new_ip']['cases'] - stats['new_ip']['pass']}</p>"
        "category_chart.png<br>"
        "<h3>IP Type Statistics</h3>"
        "type_chart.png<br>"
        "<h3>Download Excel Report</h3>"
        f"<p>{os.path.basename(excel_file)}Click here to download the Excel report</a></p>"
        "</body></html>"
    )
    html_file = os.path.join(results_dir, 'email_summary.html')
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(local_html)

    # Email HTML (CID inline images)
    email_html = (
        "<html><body>"
        "<h2>CRR Test Analysis Summary</h2>"
        "<h3>Overall Statistics</h3>"
        f"<p>Total Cases: {total_cases}<br>"
        f"   PASS: {stats['total']['pass']}<br>"
        f"   FAIL: {stats['total']['cases'] - stats['total']['pass']}<br>"
        f"   Pass Rate: {pass_rate:.1f}%</p>"
        "<id:overall_chart<br>"
        "<h3>Category Statistics</h3>"
        f"<p>Reuse IP - Cases: {stats['reuse_ip']['cases']}, PASS: {stats['reuse_ip']['pass']}, "
        f"FAIL: {stats['reuse_ip']['cases'] - stats['reuse_ip']['pass']}<br>"
        f"New IP - Cases: {stats['new_ip']['cases']}, PASS: {stats['new_ip']['pass']}, "
        f"FAIL: {stats['new_ip']['cases'] - stats['new_ip']['pass']}</p>"
        "cid:category_chart<br>"
        "<h3>IP Type Statistics</h3>"
        "cid:type_chart<br>"
        "<p>Excel report is attached.</p>"
        "</body></html>"
    )

    inline_images = [
        {"cid": "overall_chart", "path": overall_png},
        {"cid": "category_chart", "path": category_png},
        {"cid": "type_chart", "path": type_png},
    ]

    # Log paths for debugging
    print("Artifacts:")
    print(f"  HTML : {os.path.abspath(html_file)}")
    print(f"  Chart: {os.path.abspath(overall_png)}")
    print(f"  Chart: {os.path.abspath(category_png)}")
    print(f"  Chart: {os.path.abspath(type_png)}")

    return html_file, email_html, inline_images

# =========================================================
# 6) Email via NXP internal SMTP (no TLS/login)
# =========================================================
def send_email_nxp(email_html, inline_images, excel_file, recipient_email, send_flag=False):
    """
    Send HTML email with CID inline images and Excel attachment via NXP SMTP.
    SMTP host: smtp.eu-rdc02.nxp.com (no TLS, no auth, default port 25).
    """
    smtp_host = os.getenv('SMTP_HOST', 'smtp.eu-rdc02.nxp.com')
    smtp_port = int(os.getenv('SMTP_PORT', '25'))
    sender_email = os.getenv('SENDER_EMAIL', 'Dapeng <dapeng@nxp.com>')

    # Build 'related' message (HTML + inline images)
    msg_root = MIMEMultipart('related')
    msg_root['From'] = sender_email
    msg_root['To'] = recipient_email
    msg_root['Subject'] = 'CRR Test Analysis Report'

    # Alternative section: plain + HTML
    msg_alt = MIMEMultipart('alternative')
    msg_root.attach(msg_alt)
    msg_alt.attach(MIMEText('CRR Test Analysis Report - please view HTML version.', 'plain', 'utf-8'))
    msg_alt.attach(MIMEText(email_html, 'html', 'utf-8'))

    # Attach inline images
    for img in inline_images:
        with open(img['path'], 'rb') as f:
            mime_img = MIMEImage(f.read())
        mime_img.add_header('Content-ID', f"<{img['cid']}>")
        mime_img.add_header('Content-Disposition', 'inline', filename=os.path.basename(img['path']))
        msg_root.attach(mime_img)

    # Attach Excel
    with open(excel_file, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(excel_file)}"')
        msg_root.attach(part)

    if not send_flag:
        print("Email sending skipped (use --send to send). MIME preview:")
        preview_path = os.path.join('results', 'CRR_report_preview.eml')
        with open(preview_path, 'w', encoding='utf-8') as emlf:
            emlf.write(msg_root.as_string())
        print(f"EML preview saved: {os.path.abspath(preview_path)}")
        return

    # Send via NXP SMTP (no TLS/login)
    try:
        # s = smtplib.SMTP(smtp_host, timeout=15)
        # s = smtplib.SMTP("smtp.eu-rdc02.nxp.com")
        s = smtplib.SMTP("smtp.ap-rdc01.nxp.com")

        # s.ehlo()  # 可以省略；某些内部 SMTP 不要求
        s.sendmail(sender_email, [recipient_email], msg_root.as_string())
        s.quit()
        print(f"Email sent successfully to {recipient_email} via {smtp_host}:{smtp_port}")
    except Exception as e:
        print(f"[ERROR] SMTP send failed: {e}")
        fail_path = os.path.join('results', 'CRR_report_failed.eml')
        with open(fail_path, 'w', encoding='utf-8') as emlf:
            emlf.write(msg_root.as_string())
        print(f"Saved EML for manual send: {os.path.abspath(fail_path)}")

# =========================================================
# 7) Main
# =========================================================
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 CRR_test_analysis_email_nxp.py <csv_file_path> <log_directory> [--send]")
        sys.exit(1)

    csv_path = sys.argv[1]
    log_dir = sys.argv[2]
    send_flag = ('--send' in sys.argv)

    yaml_path = generate_yaml_from_csv(csv_path, log_dir)
    config = load_config(yaml_path)

    excel_file, html_file, email_html, inline_images = generate_excel_report(config)

    # 固定收件人
    recipient = 'yong.xiong@nxp.com'
    send_email_nxp(email_html, inline_images, excel_file, recipient, send_flag)
