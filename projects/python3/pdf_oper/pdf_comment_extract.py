
# import fitz  # PyMuPDF
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill
# from openpyxl.utils import get_column_letter
# from openpyxl.chart import BarChart, Reference
# # from openpyxl.chart.label import DataLabelList  # 若需要显示数据标签，可解除注释
# import os
# import argparse
# from glob import glob

# def norm(dt: str) -> str:
#     """将 PDF 的日期字符串规范化为可读格式"""
#     if not dt:
#         return ""
#     try:
#         if dt.startswith("D:"):
#             dt = dt[2:]
#         base = dt[:14]
#         from datetime import datetime
#         d = datetime.strptime(base, "%Y%m%d%H%M%S")
#         return d.isoformat(sep=" ")
#     except Exception:
#         return dt

# def get_markup_text(annot, page) -> str:
#     """提取高亮/下划线标注覆盖的原文文本"""
#     try:
#         quads = annot.vertices
#         texts = []
#         for q in quads:
#             rect = fitz.Quad(q).rect
#             texts.append(page.get_textbox(rect))
#         return " ".join(s.strip() for s in texts if s)
#     except Exception:
#         return ""

# def extract_comments(pdf_path: str):
#     """从 PDF 中提取注释/评论信息 -> list of rows"""
#     doc = fitz.open(pdf_path)
#     rows = []
#     for page_index in range(doc.page_count):
#         page = doc.load_page(page_index)
#         annots = page.annots()
#         if not annots:
#             continue
#         for annot in annots:
#             info = annot.info
#             subtype = annot.type[1]  # 'Text', 'Highlight', ...
#             content = info.get("content", "") or ""
#             author = info.get("title", "") or ""
#             created = norm(info.get("creationDate", ""))
#             modified = norm(info.get("modDate", ""))
#             rect = annot.rect
#             bbox = f"{rect.x0:.2f},{rect.y0:.2f},{rect.x1:.2f},{rect.y1:.2f}"

#             marked_text = ""
#             if subtype in ("Highlight", "Underline", "Squiggly", "StrikeOut"):
#                 marked_text = get_markup_text(annot, page)

#             rows.append([
#                 page_index + 1,
#                 subtype,
#                 author,
#                 created,
#                 modified,
#                 bbox,
#                 content,
#                 marked_text
#             ])
#     return rows

# def format_sheet(ws):
#     """通用格式：表头加粗+金色背景、自动列宽、冻结首行"""
#     header_font = Font(bold=True)
#     header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
#     for cell in ws[1]:
#         cell.font = header_font
#         cell.fill = header_fill

#     # 自动列宽
#     for col in ws.columns:
#         max_len = 0
#         for cell in col:
#             val = "" if cell.value is None else str(cell.value)
#             max_len = max(max_len, len(val))
#         ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

#     # 冻结首行 & 自动筛选
#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

# def write_formatted_excel(rows, xlsx_path: str):
#     """
#     保存为带格式的 Excel：
#     - Sheet1: All Comments（明细）
#     - Sheet2: Author Stats（作者评论数量统计 + Share 占比 + 柱状图）
#     """
#     # 明细 DataFrame
#     df = pd.DataFrame(
#         rows,
#         columns=["Page", "Type", "Author", "Created", "Modified", "BBox", "Comment", "Marked Text"]
#     )

#     # 统计 DataFrame（即使空表也生成结构）
#     if df.empty:
#         stats_df = pd.DataFrame(columns=["Author", "Count", "Share"])
#     else:
#         stats_df = (
#             df.groupby("Author", dropna=False)
#               .size()
#               .reset_index(name="Count")
#               .sort_values("Count", ascending=False)
#         )
#         total = stats_df["Count"].sum()
#         stats_df["Share"] = stats_df["Count"] / total if total else 0

#     # 写入两个工作表
#     with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
#         df.to_excel(writer, sheet_name="All Comments", index=False)
#         stats_df.to_excel(writer, sheet_name="Author Stats", index=False)

#     # 打开并加格式
#     wb = load_workbook(xlsx_path)

#     # 格式化明细
#     ws_detail = wb["All Comments"]
#     format_sheet(ws_detail)

#     # 格式化统计
#     ws_stats = wb["Author Stats"]
#     format_sheet(ws_stats)

#     # 将 Share 列设置为百分比显示
#     # 找到列索引（假设表头在第1行：A=Author, B=Count, C=Share）
#     # 若列顺序有变，可根据标题查找列号
#     share_col_letter = None
#     for cell in ws_stats[1]:
#         if str(cell.value).strip().lower() == "share":
#             share_col_letter = get_column_letter(cell.column)
#             break
#     if share_col_letter:
#         for row in range(2, ws_stats.max_row + 1):
#             ws_stats[f"{share_col_letter}{row}"].number_format = "0.00%"

#     # 添加柱状图（Author vs Count）
#     # 仅在有数据时绘制
#     if ws_stats.max_row >= 2:
#         chart = BarChart()
#         chart.title = "Comments per Author"
#         chart.style = 10
#         chart.y_axis.title = "Count"
#         chart.x_axis.title = "Author"

#         # 类别（作者）
#         cats = Reference(ws_stats, min_col=1, min_row=2, max_row=ws_stats.max_row)  # A2:A{n}
#         # 数据（数量）
#         data = Reference(ws_stats, min_col=2, min_row=1, max_row=ws_stats.max_row)  # B1:B{n} 包含表头
#         chart.add_data(data, titles_from_data=True)
#         chart.set_categories(cats)

#         # 若需要显示数据标签，可启用（部分 Excel 版本支持）
#         # from openpyxl.chart.label import DataLabelList
#         # chart.dataLabels = DataLabelList()
#         # chart.dataLabels.showVal = True

#         # 将图表插入到统计工作表中（例如放在 E2 位置）
#         ws_stats.add_chart(chart, "E2")

#     wb.save(xlsx_path)

# def process_one_pdf(pdf_path: str, outdir: str = None):
#     """处理单个 PDF 并输出 Excel"""
#     if not os.path.isfile(pdf_path):
#         print(f"跳过：不是有效文件 → {pdf_path}")
#         return

#     rows = extract_comments(pdf_path)
#     base = os.path.splitext(os.path.basename(pdf_path))[0]
#     if outdir:
#         os.makedirs(outdir, exist_ok=True)
#         xlsx_path = os.path.join(outdir, f"{base}_comments.xlsx")
#     else:
#         xlsx_path = os.path.join(os.path.dirname(pdf_path), f"{base}_comments.xlsx")

#     write_formatted_excel(rows, xlsx_path)
#     print(f"[{os.path.basename(pdf_path)}] 提取 {len(rows)} 条注释 → {xlsx_path}")

# def main():
#     parser = argparse.ArgumentParser(
#         description="从 PDF 提取注释/评论并生成带格式的 Excel（含作者统计 Sheet、占比列与柱状图）"
#     )
#     parser.add_argument(
#         "inputs",
#         nargs="+",
#         help="输入 PDF 路径（可多个），或通配符（例如：*.pdf）"
#     )
#     parser.add_argument(
#         "--outdir",
#         help="统一输出目录（可选）；默认输出到各 PDF 同目录"
#     )
#     args = parser.parse_args()

#     # 展开通配符
#     pdf_list = []
#     for p in args.inputs:
#         expanded = glob(p)
#         if expanded:
#             pdf_list.extend(expanded)
#         else:
#             pdf_list.append(p)  # 当作具体路径

#     if not pdf_list:
#         print("未找到任何 PDF 文件。")
#         return

#     for pdf_path in pdf_list:
#         process_one_pdf(pdf_path, args.outdir)

# if __name__ == "__main__":
#     main()


import os
import argparse
from glob import glob
from datetime import datetime

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ---------------------------
# 工具函数
# ---------------------------

def norm(dt: str) -> str:
    """将 PDF 的日期字符串规范化为可读格式"""
    if not dt:
        return ""
    try:
        if dt.startswith("D:"):
            dt = dt[2:]
        base = dt[:14]
        d = datetime.strptime(base, "%Y%m%d%H%M%S")
        return d.isoformat(sep=" ")
    except Exception:
        return dt or ""


def get_markup_text(annot, page) -> str:
    """提取高亮/下划线标注覆盖的原文文本"""
    try:
        quads = annot.vertices
        texts = []
        for q in quads:
            rect = fitz.Quad(q).rect
            texts.append(page.get_textbox(rect))
        return " ".join(s.strip() for s in texts if s)
    except Exception:
        return ""


# ---------------------------
# 主流程
# ---------------------------

def extract_comments(pdf_path: str):
    """提取 PDF 注释（不含 BBox）"""
    doc = fitz.open(pdf_path)
    rows = []
    for page_index in range(doc.page_count):
        page = doc.load_page(page_index)
        annots = page.annots()
        if not annots:
            continue
        for annot in annots:
            info = annot.info or {}
            subtype = annot.type[1]
            content = (info.get("content") or "").strip()
            author = (info.get("title") or "").strip() or "Unknown"
            created = norm(info.get("creationDate", ""))
            modified = norm(info.get("modDate", ""))

            marked_text = ""
            if subtype in ("Highlight", "Underline", "Squiggly", "StrikeOut"):
                marked_text = get_markup_text(annot, page)

            rows.append([
                page_index + 1,
                subtype,
                author,
                created,
                modified,
                content,
                marked_text
            ])
    return rows


def format_sheet(ws):
    """表头加粗+金色背景、自动列宽、冻结首行、自动筛选"""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_formatted_excel(rows, xlsx_path: str):
    """生成 Excel 文件（明细 + 作者统计 + 柱状图）"""
    df = pd.DataFrame(rows, columns=["Page", "Type", "Author", "Created", "Modified", "Comment", "Marked Text"])

    if df.empty:
        stats_df = pd.DataFrame(columns=["Author", "Count", "Share"])
    else:
        stats_df = df.groupby("Author", dropna=False).size().reset_index(name="Count").sort_values("Count", ascending=False)
        total = stats_df["Count"].sum()
        stats_df["Share"] = stats_df["Count"] / total if total else 0

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Comments", index=False)
        stats_df.to_excel(writer, sheet_name="Author Stats", index=False)

    wb = load_workbook(xlsx_path)
    ws_detail = wb["All Comments"]
    ws_stats = wb["Author Stats"]

    format_sheet(ws_detail)
    format_sheet(ws_stats)

    # Share 列百分比格式
    for cell in ws_stats[1]:
        if str(cell.value).lower() == "share":
            col_letter = get_column_letter(cell.column)
            for row in range(2, ws_stats.max_row + 1):
                ws_stats[f"{col_letter}{row}"].number_format = "0.00%"

    # 柱状图
    if ws_stats.max_row >= 2:
        chart = BarChart()
        chart.title = "Comments per Author"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Author"
        cats = Reference(ws_stats, min_col=1, min_row=2, max_row=ws_stats.max_row)
        data = Reference(ws_stats, min_col=2, min_row=1, max_row=ws_stats.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_stats.add_chart(chart, "E2")

    wb.save(xlsx_path)


def process_one_pdf(pdf_path: str, outdir: str = None):
    """处理单个 PDF 并输出 Excel，文件名带年月日时分"""
    if not os.path.isfile(pdf_path):
        print(f"跳过：不是有效文件 → {pdf_path}")
        return

    rows = extract_comments(pdf_path)
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    date_str = datetime.now().strftime("%Y%m%d_%H%M")  # 年月日_时分

    if outdir:
        os.makedirs(outdir, exist_ok=True)
        xlsx_path = os.path.join(outdir, f"{base}_comments_{date_str}.xlsx")
    else:
        xlsx_path = os.path.join(os.path.dirname(pdf_path), f"{base}_comments_{date_str}.xlsx")

    write_formatted_excel(rows, xlsx_path)
    print(f"[{os.path.basename(pdf_path)}] 提取 {len(rows)} 条注释 → {xlsx_path}")


def main():
    parser = argparse.ArgumentParser(description="提取 PDF 注释并生成 Excel（文件名带年月日时分）")
    parser.add_argument("inputs", nargs="+", help="输入 PDF 路径（可多个或通配符）")
    parser.add_argument("--outdir", help="统一输出目录（可选）")
    args = parser.parse_args()

    pdf_list = []
    for p in args.inputs:
        expanded = glob(p)
        pdf_list.extend(expanded if expanded else [p])

    if not pdf_list:
        print("未找到任何 PDF 文件。")
        return

    for pdf_path in pdf_list:
        process_one_pdf(pdf_path, args.outdir)


if __name__ == "__main__":
    main()


