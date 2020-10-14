# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name:     got_house_info
   Description: 
   Author:        nxa28190
   Date:          7/5/2020
-------------------------------------------------
   Change Activity:
                  7/5/2020:
-------------------------------------------------
"""
__author__ = 'Richard Xiong'

import requests
import time
import sys
from selenium import webdriver
from selenium.webdriver.support.select import Select
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import YELLOW
from datetime import datetime


class ExcelOper(object):
    def __init__(self):
        self.wb = Workbook()
        self.rename_first_sheet()

    def rename_first_sheet(self):
        # set the first default sheet name as Summary
        first_sheet = self.wb.active
        first_sheet.title = 'First'

    def create_sheet(self, name, index=None):
        self.wb.create_sheet(title=name, index=index)

    def set_value(self, sheet_name, row, column, value, color):
        # self.create_sheet(sheet_name)
        ws = self.wb[sheet_name]
        # ws.cell(row=row, column=column).value = value
        ws['{}{}'.format(column, row)].value = value

        # set alignment
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['{}{}'.format(column, row)].alignment = align

        # set backend color
        if color == 'green':
            fill = PatternFill("solid", fgColor='66cc33')
        elif color == 'grey':
            fill = PatternFill("solid", fgColor='cccccc')
        elif color == 'yellow':
            fill = PatternFill("solid", fgColor=YELLOW)
        elif color == 'brown':
            fill = PatternFill("solid", fgColor='666600')
        ws['{}{}'.format(column, row)].fill = fill

        return ws

    def set_column_width(self, sheet_name, length, column='B'):
        ws = self.wb.get_sheet_by_name(sheet_name)
        ws.column_dimensions[column].width = length
        return ws

    def set_alignment(self, sheet_name, row, column, hstyle='center', vstyle='center', wrap_text=True):
        ws = self.wb.get_sheet_by_name(sheet_name)
        align = Alignment(horizontal=hstyle, vertical=vstyle, wrap_text=wrap_text)
        ws['{}{}'.format(column, row)].alignment = align
        return ws

    def set_font(self, sheet_name, row, column, name='Calibri', bold=False, color='FF000000'):
        ws = self.wb.get_sheet_by_name(sheet_name)
        font = Font(name=name, bold=bold, color=color)
        ws.cell(row=row, column=column).font = font
        return ws

    def set_hyper_link(self, sheet_name, row, column, hyper_sheet_name):
        ws = self.wb.get_sheet_by_name(sheet_name)
        ws.cell(row=row, column=column).value = hyper_sheet_name
        ws.cell(row=row, column=column).hyperlink = '#{}!A1'.format(hyper_sheet_name)
        return ws

    def active_sheet(self, sheet_name):
        ws = self.wb.get_sheet_by_name(sheet_name)
        return ws

    def save_wb(self, excel_name):
        self.wb.save(excel_name)


class HouseInfo(ExcelOper):
    def __init__(self, project, region):
        self.chrome = webdriver.Chrome()
        # self.chrome = webdriver.Ie()
        super(HouseInfo, self).__init__()

        self.chrome.get(BASE_URL)
        # input house info
        self.chrome.find_element_by_xpath('// *[ @ id = "ctl00_MainContent_txt_Pro"]').send_keys(project)

        # select region
        house_select = Select(self.chrome.find_element_by_id('ctl00_MainContent_ddl_RD_CODE'))
        house_select.select_by_value(region)
        # click search button
        self.chrome.find_element_by_xpath('//*[@id="ctl00_MainContent_bt_select"]').click()

    def get_house_set_lnk(self):
        # got the url link to count the house set number
        page_tree = etree.HTML(house.chrome.page_source)
        house_set = page_tree.xpath('//*[@id="ctl00_MainContent_OraclePager1"]/tbody/child::tr/child::*/a/@href')
        house_set_link = []
        for i in range(1, len(house_set) + 1, 1):
            house_set_link.append('//*[@id="ctl00_MainContent_OraclePager1"]/tbody/tr[{}]/td/a'.format(i + 1))
        return house_set_link

    def enter_house_set(self, url):
        # enter first house set
        self.chrome.find_element_by_xpath(url).click()
        time.sleep(0.5)

    def get_house_link(self):
        # got the url link to count the house set number
        page_tree = etree.HTML(house.chrome.page_source)
        building_no = page_tree.xpath('//*[@id="ctl00_MainContent_OraclePager1"]/tbody/child::tr/child::*/a/text()')
        print(building_no)
        house_link = []
        for i in range(1, len(building_no)+1, 1):
            house_link.append('//*[@id="ctl00_MainContent_OraclePager1"]/tbody/tr[{}]/td/a'.format(i + 1))
        # add all the build number and link to a list
        building_info = []
        for no, link in zip(building_no, house_link):
            info = {
                'number': no,
                'link': link
            }
            building_info.append(info)
        return building_info

    def back(self):
        self.chrome.back()

    def enter_house_info_page(self, house_link):
        # enter house derailed information page
        self.chrome.find_element_by_xpath(house_link).click()
        time.sleep(1)

        # got the house color and number
        page_tree = etree.HTML(self.chrome.page_source)
        house_no = page_tree.xpath('//*[@id="ctl00_MainContent_gvxml"]/tbody/child::tr/child::*/text()')
        house_color = page_tree.xpath('//*[@id="ctl00_MainContent_gvxml"]/tbody/child::tr/child::*/@style')

        print(house_color)
        print(house_no)

        for item in house_no[:]:
            if ('\xa0' == item) or ('地下车库' == item):
                print(item)
                house_no.remove(item)

        for item in house_color[:]:
            if ('background-color:White;' == item):
                print(item)
                house_color.remove(item)

        house_info = []
        for number, color in zip(house_no, house_color):
            info = {
                'number': number,
                'color': color
            }
            house_info.append(info)

        # remove the blank house information
        not_sale = 0
        for item in house_info:
            # if (('\xa0' or '地下车库') in item['number']) or ('background-color:White;' in item['color']):
            #     print(item)
            #     house_info.remove(item)
            if 'background-color:#66cc33;' in item['color']:
                not_sale = not_sale + 1

        all_house = len(house_info)
        sale_rate = (all_house - not_sale)/all_house

        print(house_color, house_no)
        print(house_info)
        print(not_sale)
        print(all_house)
        print(sale_rate)

        self.chrome.back()
        return not_sale, all_house, sale_rate, house_info


if __name__ == '__main__':
    COLUM_VALUE = ['A', 'B', 'C', 'D', 'E']
    COLOR = {'background-color:#66cc33;': 'green', 'background-color:#cccccc;': 'grey',
             'background-color:Yellow;': 'yellow', 'background-color:#666600;': 'brown'}
    BASE_URL = r'http://spf.szfcweb.com/szfcweb/(S(wwkvmt45ppfml2mkbrlotu35))/DataSerach/SaleInfoProListIndex.aspx'

    house = HouseInfo(sys.argv[1], sys.argv[2])
    # house = HouseInfo('都会', '吴中区')
    # house_set_link = house.get_house_set_lnk()

    # add for storing sales information to first page
    colum_index = ['A','B','C','D','E','F','G','H','I','J','K','L']
    i = 0
    for house_set_link in house.get_house_set_lnk():
        house.chrome.find_element_by_xpath(house_set_link).click()

        for building in house.get_house_link():
            # house.chrome.find_element_by_xpath(building['link']).click()

            house.create_sheet(str(building['number']))
            not_sale, all_house, sale_rate, house_info = house.enter_house_info_page(building['link'])
            for number, info in enumerate(house_info):
                row_value = int(number / 4 + 1)
                colum_value = number % 4
                house.set_value(sheet_name=str(building['number']), row=row_value, column=COLUM_VALUE[colum_value],
                                value=info['number'], color=COLOR[info['color']])

            house.set_value(sheet_name=str(building['number']), row=1, column='F', value='去化率', color='green')
            house.set_value(sheet_name=str(building['number']), row=2, column='F', value=sale_rate, color='green')
            house.set_value(sheet_name=str(building['number']), row=4, column='F', value='未售', color='green')
            house.set_value(sheet_name=str(building['number']), row=5, column='F', value=not_sale, color='green')
            house.set_value(sheet_name=str(building['number']), row=4, column='G', value='总量', color='green')
            house.set_value(sheet_name=str(building['number']), row=5, column='G', value=all_house, color='green')
            house.set_value(sheet_name=str(building['number']), row=6, column='F', value='已售', color='green')
            house.set_value(sheet_name=str(building['number']), row=7, column='F', value=all_house - not_sale, color='green')

            # add sales information to the first page
            house.set_value(sheet_name='First', row=1, column=colum_index[i], value=all_house - not_sale, color='green')
            i = i + 1

        house.back()

    excel_name = '{}_{}_房屋销售统计{}.xlsx'.format(sys.argv[1], sys.argv[2], datetime.now().strftime('%y_%m_%d-%H-%M'))
    print(excel_name)
    house.save_wb(excel_name)

